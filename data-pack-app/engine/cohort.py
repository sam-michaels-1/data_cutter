"""
Cohort tab generators.

Generates Annual and Quarterly cohort analysis tabs.

Each tab has multiple blocks:
  Block 1: "Total Business" (all filters = "<>")
  Block 2+: Filtered blocks

Each block has 4 horizontal sections:
  1. ARR by period
  2. Customer count by period
  3. ARR Retention (indexed to starting period)
  4. Logo Retention (indexed to starting period)

Reference structure (Annual Cohort):
  Row 6: Title ("Total Business Annual ARR by Cohort")
  Row 8: Section headers ("Annual ARR", "Annual Customers", etc.)
  Row 9: Column headers (B=Quarter, C=Year, D-F=filters, G=Cohort, H+=dates)
  Row 10+: Cohort data rows
  After data: Total, Average, Median, Weighted Average, Check
"""
from openpyxl.utils import get_column_letter
from .utils import col_letter, get_yoy_offset


def generate_cohort_tab(wb, config, clean_sheet_name, clean_layout,
                        first_data_row, last_data_row, granularity,
                        filter_blocks):
    """
    Generate a cohort analysis tab.

    granularity: 'annual' or 'quarterly' - determines the cohort period
    """
    sheet_name = f"{granularity.capitalize()} Cohort"
    ws = wb.create_sheet(sheet_name)

    num_dates = clean_layout['num_dates']
    num_attrs = clean_layout['num_attrs']
    attr_names = list(config['attributes'].keys())
    num_cohorts = num_dates  # one cohort per period

    # Layout columns
    # B: Quarter, C: Year, D-F: filters, G: cohort label
    q_col = 2   # B
    y_col = 3   # C
    filter_start = 4  # D
    filter_end = filter_start + num_attrs - 1
    cohort_label_col = filter_end + 1  # G

    # Section 1: ARR (one column per period)
    s1_start = cohort_label_col + 1  # H
    s1_end = s1_start + num_dates - 1

    # Section 2: Customer count
    s2_start = s1_end + 2  # gap + start
    s2_end = s2_start + num_dates - 1

    # Section 3: ARR Retention
    s3_label_col = s2_end + 2
    s3_start_val_col = s3_label_col + 1  # "Starting ARR/Size" column
    s3_data_start = s3_start_val_col + 1
    s3_data_end = s3_data_start + num_dates - 1

    # Section 4: Logo Retention
    s4_label_col = s3_data_end + 2
    s4_start_val_col = s4_label_col + 1
    s4_data_start = s4_start_val_col + 1
    s4_data_end = s4_data_start + num_dates - 1

    # Counter row (row 1) for retention period labels
    # S3 gets 0,1,2,... and S4 references S3
    ws.cell(row=1, column=s3_data_start, value=0)
    for i in range(1, num_dates):
        prev3 = col_letter(s3_data_start + i - 1)
        ws.cell(row=1, column=s3_data_start + i, value=f"={prev3}1+1")
    # S4 counters reference S3 counters
    for i in range(num_dates):
        ws.cell(row=1, column=s4_data_start + i,
                value=f"={col_letter(s3_data_start + i)}1")

    # Check summary
    ws.cell(row=1, column=1, value="Check Summary")

    # Units (A3=label, B3=formula)
    ws.cell(row=3, column=1, value="Units")
    ws.cell(row=3, column=q_col, value="=Control!$C$4")
    units_cell = f"${col_letter(q_col)}$3"

    # Period prefix for labels
    if granularity == 'quarterly':
        period_prefix = 'Q'
    else:
        period_prefix = 'Y'

    check_refs = []

    for block_idx, block in enumerate(filter_blocks):
        block_start = 6 + block_idx * (num_cohorts + 9)

        _write_cohort_block(
            ws, block_start, block, config,
            clean_sheet_name, clean_layout,
            first_data_row, last_data_row,
            q_col, y_col, filter_start, filter_end, cohort_label_col,
            s1_start, s1_end, s2_start, s2_end,
            s3_label_col, s3_start_val_col, s3_data_start, s3_data_end,
            s4_label_col, s4_start_val_col, s4_data_start, s4_data_end,
            num_dates, num_cohorts, num_attrs, attr_names,
            units_cell, granularity, period_prefix
        )

        # Collect check row references
        check_row = block_start + 3 + num_cohorts + 4
        s1s = col_letter(s1_start)
        s1e = col_letter(s1_end)
        s2s = col_letter(s2_start)
        s2e = col_letter(s2_end)
        check_refs.append(f"{s1s}{check_row}:{s1e}{check_row}")
        check_refs.append(f"{s2s}{check_row}:{s2e}{check_row}")

    # Write check summary
    if check_refs:
        ws.cell(row=1, column=2, value=f"=SUM({','.join(check_refs)})")

    return sheet_name


def _write_cohort_block(ws, start_row, block, config,
                        clean_sheet, clean_layout,
                        cdr_first, cdr_last,
                        q_col, y_col, filter_start, filter_end, cohort_label_col,
                        s1_start, s1_end, s2_start, s2_end,
                        s3_label, s3_start_val, s3_data_start, s3_data_end,
                        s4_label, s4_start_val, s4_data_start, s4_data_end,
                        num_dates, num_cohorts, num_attrs, attr_names,
                        units_cell, granularity, period_prefix):
    """Write a single cohort analysis block."""
    title = block['title']
    filters = block['filters']

    r_title = start_row
    r_section_headers = start_row + 2  # section headers row
    r_headers = start_row + 3  # column headers row

    # Title (at B column)
    metric_label = "ARR" if config.get("data_type", "arr") == "arr" else "Revenue"
    ws.cell(row=r_title, column=q_col,
            value=f"{title} {granularity.capitalize()} {metric_label} by Cohort")

    # Section headers
    ws.cell(row=r_section_headers, column=s1_start,
            value=f"{granularity.capitalize()} {metric_label}")
    ws.cell(row=r_section_headers, column=s2_start,
            value=f"{granularity.capitalize()} Customers")
    ws.cell(row=r_section_headers, column=s3_label,
            value=f"{granularity.capitalize()} {metric_label} Retention")
    ws.cell(row=r_section_headers, column=s4_label,
            value=f"{granularity.capitalize()} Logo Retention")

    # Column headers row
    ws.cell(row=r_headers, column=q_col, value="Quarter")
    ws.cell(row=r_headers, column=y_col, value="Year")
    for attr_idx, attr_name in enumerate(attr_names):
        ws.cell(row=r_headers, column=filter_start + attr_idx, value=attr_name)
    ws.cell(row=r_headers, column=cohort_label_col, value="Cohort")

    # Period headers - reference clean data
    for i in range(num_dates):
        clean_date_col = clean_layout['arr_start'] + i
        ws.cell(row=r_headers, column=s1_start + i,
                value=f"='{clean_sheet}'!{col_letter(clean_date_col)}$6")
        ws.cell(row=r_headers, column=s2_start + i,
                value=f"={col_letter(s1_start + i)}{r_headers}")

    # Retention section headers
    ws.cell(row=r_headers, column=s3_label, value="Cohort")
    ws.cell(row=r_headers, column=s3_start_val, value="Starting Size")
    ws.cell(row=r_headers, column=s4_label, value="Cohort")
    ws.cell(row=r_headers, column=s4_start_val, value="Starting Size")

    # Retention period headers (Y0, Y1, ... or Q0, Q1, ...)
    for i in range(num_dates):
        ws.cell(row=r_headers, column=s3_data_start + i,
                value=f'="{period_prefix}"&{col_letter(s3_data_start + i)}$1')
        ws.cell(row=r_headers, column=s4_data_start + i,
                value=f'="{period_prefix}"&{col_letter(s4_data_start + i)}$1')

    # First cohort row
    first_cohort_row = r_headers + 1

    # Source column references
    src_arr_start = col_letter(clean_layout['arr_start'])
    src_arr_end = col_letter(clean_layout['arr_end'])

    # Write cohort data rows
    for cohort_idx in range(num_cohorts):
        row = first_cohort_row + cohort_idx

        # Quarter column
        if cohort_idx == 0:
            ws.cell(row=row, column=q_col,
                    value=f"='{clean_sheet}'!${src_arr_start}$2")
        else:
            prev_q = col_letter(q_col)
            if granularity == 'quarterly':
                ws.cell(row=row, column=q_col,
                        value=f"=IF({prev_q}{row-1}=4,1,{prev_q}{row-1}+1)")
            else:
                ws.cell(row=row, column=q_col,
                        value=f"={prev_q}{row-1}")

        # Year column
        if cohort_idx == 0:
            ws.cell(row=row, column=y_col,
                    value=f"='{clean_sheet}'!${src_arr_start}$3")
        else:
            prev_y = col_letter(y_col)
            prev_q = col_letter(q_col)
            if granularity == 'quarterly':
                ws.cell(row=row, column=y_col,
                        value=f"=IF({prev_q}{row-1}=4,{prev_y}{row-1}+1,{prev_y}{row-1})")
            else:
                ws.cell(row=row, column=y_col,
                        value=f"={prev_y}{row-1}+1")

        # Filter columns
        for attr_idx, attr_name in enumerate(attr_names):
            fc = filter_start + attr_idx
            if cohort_idx == 0:
                ws.cell(row=row, column=fc,
                        value=filters.get(attr_name, "<>"))
            else:
                ws.cell(row=row, column=fc,
                        value=f"={col_letter(fc)}{row-1}")

        # Cohort label
        ql = col_letter(q_col)
        yl = col_letter(y_col)
        if granularity == 'quarterly':
            ws.cell(row=row, column=cohort_label_col,
                    value=f'="Q"&{ql}{row}&"\'"&RIGHT({yl}{row},2)')
        else:
            ws.cell(row=row, column=cohort_label_col,
                    value=f'="FY"&"\'"&RIGHT({yl}{row},2)')

        # Build SUMIFS criteria for this cohort
        gl = col_letter(cohort_label_col)
        criteria_parts = []
        for attr_idx in range(num_attrs):
            clean_col = clean_layout['attr_start'] + attr_idx
            criteria_parts.append(
                f"'{clean_sheet}'!${col_letter(clean_col)}${cdr_first}"
                f":${col_letter(clean_col)}${cdr_last},"
                f"${col_letter(filter_start + attr_idx)}{row}")
        # Cohort criterion
        criteria_parts.append(
            f"'{clean_sheet}'!${col_letter(clean_layout['cohort'])}${cdr_first}"
            f":${col_letter(clean_layout['cohort'])}${cdr_last},"
            f"${gl}{row}")
        criteria_str = ','.join(criteria_parts)

        # Section 1: ARR per period
        for period_idx in range(num_dates):
            clean_col = clean_layout['arr_start'] + period_idx
            sum_range = (f"'{clean_sheet}'!{col_letter(clean_col)}${cdr_first}"
                         f":{col_letter(clean_col)}${cdr_last}")
            dc = s1_start + period_idx
            ws.cell(row=row, column=dc,
                    value=f"=SUMIFS({sum_range},{criteria_str})/{units_cell}")

        # Section 2: Customer count per period
        for period_idx in range(num_dates):
            clean_col = clean_layout['arr_start'] + period_idx
            count_range = (f"'{clean_sheet}'!{col_letter(clean_col)}${cdr_first}"
                           f":{col_letter(clean_col)}${cdr_last}")
            dc = s2_start + period_idx
            ws.cell(row=row, column=dc,
                    value=f'=COUNTIFS({count_range},"<>"&0,{criteria_str})')

        # Section 3: ARR Retention
        s1s = col_letter(s1_start)
        s1e = col_letter(s1_end)
        # Cohort label
        ws.cell(row=row, column=s3_label, value=f"={gl}{row}")
        # Starting ARR (XLOOKUP to find the cohort's starting period ARR)
        s3sv = col_letter(s3_start_val)
        ws.cell(row=row, column=s3_start_val,
                value=f"=_xlfn.XLOOKUP({col_letter(s3_label)}{row},"
                      f"{col_letter(s1_start)}${r_headers}:{col_letter(s1_end)}${r_headers},"
                      f"{s1s}{row}:{s1e}{row})")

        # Retention ratios: ARR[period] / Starting ARR (left-adjusted)
        for period_idx in range(num_dates - cohort_idx):
            arr_col = s1_start + cohort_idx + period_idx
            dc = s3_data_start + period_idx
            ws.cell(row=row, column=dc,
                    value=f"={col_letter(arr_col)}{row}/${s3sv}{row}")

        # Section 4: Logo Retention
        s2s = col_letter(s2_start)
        s2e = col_letter(s2_end)
        ws.cell(row=row, column=s4_label, value=f"={gl}{row}")
        s4sv = col_letter(s4_start_val)
        ws.cell(row=row, column=s4_start_val,
                value=f"=_xlfn.XLOOKUP({col_letter(s4_label)}{row},"
                      f"{col_letter(s2_start)}${r_headers}:{col_letter(s2_end)}${r_headers},"
                      f"{s2s}{row}:{s2e}{row})")

        for period_idx in range(num_dates - cohort_idx):
            count_col = s2_start + cohort_idx + period_idx
            dc = s4_data_start + period_idx
            ws.cell(row=row, column=dc,
                    value=f"={col_letter(count_col)}{row}/${s4sv}{row}")

    # Summary rows (compact: Total+Average share a row)
    last_cohort_row = first_cohort_row + num_cohorts - 1
    r_total = last_cohort_row + 1   # S1/S2 Total + S3/S4 Average
    r_median = r_total + 1          # S3/S4 Median
    r_weighted = r_median + 1       # S3/S4 Weighted Average
    r_check = r_weighted + 1        # S1/S2 Check

    # Filter columns for summary rows
    for attr_idx in range(num_attrs):
        fc = filter_start + attr_idx
        for r in [r_total, r_median, r_weighted, r_check]:
            ws.cell(row=r, column=fc,
                    value=f"={col_letter(fc)}{r-1}")

    # Total row
    for period_idx in range(num_dates):
        dc = s1_start + period_idx
        dcl = col_letter(dc)
        ws.cell(row=r_total, column=dc,
                value=f"=SUM({dcl}{first_cohort_row}:{dcl}{last_cohort_row})")

        dc2 = s2_start + period_idx
        dc2l = col_letter(dc2)
        ws.cell(row=r_total, column=dc2,
                value=f"=SUM({dc2l}{first_cohort_row}:{dc2l}{last_cohort_row})")

    # Average, Median, Weighted for retention columns
    # Average goes on r_total row (same row as S1/S2 Total)
    for period_idx in range(num_dates):
        # Section 3 (ARR retention)
        dc3 = s3_data_start + period_idx
        dc3l = col_letter(dc3)
        ws.cell(row=r_total, column=dc3,
                value=f"=AVERAGE({dc3l}{first_cohort_row}:{dc3l}{last_cohort_row})")
        ws.cell(row=r_median, column=dc3,
                value=f"=MEDIAN({dc3l}{first_cohort_row}:{dc3l}{last_cohort_row})")

        # Dollar-weighted average
        s3svl = col_letter(s3_start_val)
        ws.cell(row=r_weighted, column=dc3,
                value=f'=SUMPRODUCT(({dc3l}{first_cohort_row}:{dc3l}{last_cohort_row}<>"")'
                      f'*{dc3l}{first_cohort_row}:{dc3l}{last_cohort_row}'
                      f',${s3svl}{first_cohort_row}:${s3svl}{last_cohort_row})'
                      f'/SUMPRODUCT(({dc3l}{first_cohort_row}:{dc3l}{last_cohort_row}<>"")*1'
                      f',${s3svl}{first_cohort_row}:${s3svl}{last_cohort_row})')

        # Section 4 (Logo retention)
        dc4 = s4_data_start + period_idx
        dc4l = col_letter(dc4)
        ws.cell(row=r_total, column=dc4,
                value=f"=AVERAGE({dc4l}{first_cohort_row}:{dc4l}{last_cohort_row})")
        ws.cell(row=r_median, column=dc4,
                value=f"=MEDIAN({dc4l}{first_cohort_row}:{dc4l}{last_cohort_row})")

        # Size-weighted average
        s4svl = col_letter(s4_start_val)
        ws.cell(row=r_weighted, column=dc4,
                value=f'=SUMPRODUCT(({dc4l}{first_cohort_row}:{dc4l}{last_cohort_row}<>"")'
                      f'*{dc4l}{first_cohort_row}:{dc4l}{last_cohort_row}'
                      f',${s4svl}{first_cohort_row}:${s4svl}{last_cohort_row})'
                      f'/SUMPRODUCT(({dc4l}{first_cohort_row}:{dc4l}{last_cohort_row}<>"")*1'
                      f',${s4svl}{first_cohort_row}:${s4svl}{last_cohort_row})')

    # Check row
    for period_idx in range(num_dates):
        # Section 1 check (ARR)
        dc = s1_start + period_idx
        dcl = col_letter(dc)
        clean_col = clean_layout['arr_start'] + period_idx
        criteria_check = []
        for attr_idx in range(num_attrs):
            cc = clean_layout['attr_start'] + attr_idx
            criteria_check.append(
                f"'{clean_sheet}'!${col_letter(cc)}${cdr_first}"
                f":${col_letter(cc)}${cdr_last},"
                f"${col_letter(filter_start + attr_idx)}{r_check}")
        crit_str = ','.join(criteria_check)
        sum_range = (f"'{clean_sheet}'!{col_letter(clean_col)}${cdr_first}"
                     f":{col_letter(clean_col)}${cdr_last}")
        if crit_str:
            ws.cell(row=r_check, column=dc,
                    value=f"={dcl}{r_total}*{units_cell}-SUMIFS({sum_range},{crit_str})")
        else:
            ws.cell(row=r_check, column=dc,
                    value=f"={dcl}{r_total}*{units_cell}-SUM({sum_range})")

        # Section 2 check (customers)
        dc2 = s2_start + period_idx
        dc2l = col_letter(dc2)
        count_range = (f"'{clean_sheet}'!{col_letter(clean_col)}${cdr_first}"
                       f":{col_letter(clean_col)}${cdr_last}")
        if crit_str:
            ws.cell(row=r_check, column=dc2,
                    value=f'={dc2l}{r_total}-COUNTIFS({count_range},"<>"&0,{crit_str})')
        else:
            ws.cell(row=r_check, column=dc2,
                    value=f'={dc2l}{r_total}-COUNTIF({count_range},"<>"&0)')

    # Row labels for summary
    ws.cell(row=r_total, column=cohort_label_col, value="Total")
    ws.cell(row=r_total, column=s3_label, value="Average")
    ws.cell(row=r_total, column=s4_label, value="Average")
    ws.cell(row=r_median, column=s3_label, value="Median")
    ws.cell(row=r_median, column=s4_label, value="Median")
    ws.cell(row=r_weighted, column=s3_label, value="Dollar-Weighted Average")
    ws.cell(row=r_weighted, column=s4_label, value="Size-Weighted Average")
    ws.cell(row=r_check, column=cohort_label_col, value="Check")
