"""
Main orchestrator for the Data Pack Excel generator.

Takes a mapping configuration and a raw Excel file, produces the
fully formula-driven output workbook.
"""
import openpyxl
from collections import OrderedDict
from .clean_data import generate_base_clean_data, generate_aggregated_clean_data
from .retention import generate_retention_tab
from .cohort import generate_cohort_tab
from .top_customers import generate_top_customers_tab, TOP_N, CONCENTRATION_TIERS
from .utils import get_yoy_offset, compute_clean_layout
from .formatting import (
    format_control_tab, format_clean_data_tab, format_retention_tab,
    format_cohort_tab, format_top_customers_tab, apply_formula_coloring
)


def generate_data_pack(config, input_file, output_file):
    """
    Generate the complete data pack Excel file.

    config: dict with mapping configuration
    input_file: path to the raw Excel file
    output_file: path for the output Excel file
    """
    print("Reading raw data...")
    src_wb = openpyxl.load_workbook(input_file, data_only=True)
    src_ws = src_wb[config['raw_data_sheet']]

    # Extract unique dates and customers from raw data
    unique_dates, unique_customers = _extract_uniques(src_ws, config)
    print(f"  Found {len(unique_customers)} unique customers")
    print(f"  Found {len(unique_dates)} unique time periods")

    # Create output workbook
    wb = openpyxl.Workbook()

    # --- Tab 1: Control ---
    print("Creating Control tab...")
    _create_control_tab(wb, config)

    # --- Copy raw data ---
    print("Copying raw data...")
    _copy_raw_data(wb, src_wb, config)

    # --- Determine which tabs to generate based on granularity ---
    granularity = config['time_granularity']
    output_grans = config.get('output_granularities')  # None → use default
    clean_tabs = {}  # {granularity: (sheet_name, layout, first_row, last_row)}

    # --- Base clean data tab ---
    print(f"Generating Clean {granularity.capitalize()} Data...")
    base_result = generate_base_clean_data(wb, config, unique_dates, unique_customers)
    clean_tabs[granularity] = base_result

    # --- Aggregated tabs ---
    if granularity == 'monthly':
        # Generate quarterly from monthly
        quarterly_dates = _compute_quarterly_dates(unique_dates, config)
        print(f"Generating Clean Quarterly Data ({len(quarterly_dates)} quarters)...")
        q_result = generate_aggregated_clean_data(
            wb, config, base_result[0], base_result[1],
            'quarterly', unique_customers, quarterly_dates)
        clean_tabs['quarterly'] = q_result

        # Generate annual from monthly
        annual_dates = _compute_annual_dates(unique_dates, config)
        print(f"Generating Clean Annual Data ({len(annual_dates)} years)...")
        a_result = generate_aggregated_clean_data(
            wb, config, base_result[0], base_result[1],
            'annual', unique_customers, annual_dates)
        clean_tabs['annual'] = a_result

    elif granularity == 'quarterly':
        # Generate annual from quarterly
        annual_dates = _compute_annual_from_quarterly_dates(unique_dates, config)
        print(f"Generating Clean Annual Data ({len(annual_dates)} years)...")
        a_result = generate_aggregated_clean_data(
            wb, config, base_result[0], base_result[1],
            'annual', unique_customers, annual_dates)
        clean_tabs['annual'] = a_result

    # --- Build filter blocks for retention/cohort ---
    filter_blocks = _build_filter_blocks(config)
    num_attrs = len(config['attributes'])

    # --- Retention tabs ---
    for g in _get_available_granularities(granularity, output_grans):
        if g in clean_tabs:
            sheet_name, layout, fdr, ldr = clean_tabs[g]
            print(f"Generating {g.capitalize()} Retention...")
            generate_retention_tab(
                wb, config, sheet_name, layout, fdr, ldr, g, filter_blocks)

            # Compute retention layout columns for formatting
            yoy_offset = get_yoy_offset(g)
            num_derived = layout['num_dates'] - yoy_offset
            filter_start_col = 2
            cohort_fc = filter_start_col + num_attrs
            s1_label = cohort_fc + 2
            s1_start = s1_label + 1
            s1_end = s1_start + num_derived - 1
            s2_label = s1_end + 2
            s2_start = s2_label + 1
            s2_end = s2_start + num_derived - 1
            s3_label = s2_end + 2
            s3_start = s3_label + 1
            s3_end = s3_start + num_derived - 1

            ret_sheet = f"{g.capitalize()} Retention"
            print(f"  Formatting {ret_sheet}...")
            format_retention_tab(
                wb[ret_sheet], config, filter_blocks, num_derived, num_attrs,
                s1_label, s1_start, s1_end,
                s2_label, s2_start, s2_end,
                s3_label, s3_start, s3_end,
                filter_start_col, cohort_fc)

    # --- Cohort tabs ---
    for g in ['quarterly', 'annual']:
        if g in clean_tabs:
            sheet_name, layout, fdr, ldr = clean_tabs[g]
            print(f"Generating {g.capitalize()} Cohort...")
            generate_cohort_tab(
                wb, config, sheet_name, layout, fdr, ldr, g, filter_blocks)

            # Compute cohort layout columns for formatting
            num_dates = layout['num_dates']
            num_cohorts = num_dates
            q_col = 2
            y_col = 3
            filter_start = 4
            filter_end = filter_start + num_attrs - 1
            cohort_label_col = filter_end + 1
            s1_start = cohort_label_col + 1
            s1_end = s1_start + num_dates - 1
            s2_start = s1_end + 2
            s2_end = s2_start + num_dates - 1
            s3_label_col = s2_end + 2
            s3_start_val_col = s3_label_col + 1
            s3_data_start = s3_start_val_col + 1
            s3_data_end = s3_data_start + num_dates - 1
            s4_label_col = s3_data_end + 2
            s4_start_val_col = s4_label_col + 1
            s4_data_start = s4_start_val_col + 1
            s4_data_end = s4_data_start + num_dates - 1

            coh_sheet = f"{g.capitalize()} Cohort"
            print(f"  Formatting {coh_sheet}...")
            format_cohort_tab(
                wb[coh_sheet], config, filter_blocks,
                num_dates, num_cohorts, num_attrs,
                q_col, y_col, filter_start, cohort_label_col,
                s1_start, s1_end, s2_start, s2_end,
                s3_label_col, s3_start_val_col, s3_data_start, s3_data_end,
                s4_label_col, s4_start_val_col, s4_data_start, s4_data_end,
                g)

    # --- Top Customer Analysis (always based on annual if available) ---
    if 'annual' in clean_tabs:
        sheet_name, layout, fdr, ldr = clean_tabs['annual']
        print("Generating Annual Top Customer Analysis...")
        generate_top_customers_tab(wb, config, sheet_name, layout, fdr, ldr)

        # Format top customers tab
        num_dates = layout['num_dates']
        rank_num_col = 2
        cust_id_col = 3
        attr_start = 4
        cohort_col = attr_start + num_attrs
        s1_start = cohort_col + 1
        s1_end = s1_start + num_dates - 1
        s2_start = s1_end + 2
        s2_end = s2_start + num_dates - 2
        s3_start = s2_end + 2
        s3_end = s3_start + num_dates - 1
        first_customer_row = 7
        last_customer_row = first_customer_row + TOP_N - 1
        r_top_total = last_customer_row + 1
        r_other = r_top_total + 1
        r_total = r_other + 1
        r_memo_start = r_total + 2

        print("  Formatting Annual Top Customer Analysis...")
        format_top_customers_tab(
            wb['Annual Top Customer Analysis'], config, layout,
            first_customer_row, last_customer_row,
            r_top_total, r_other, r_total, r_memo_start,
            num_dates,
            rank_num_col, cust_id_col, attr_start, num_attrs, cohort_col,
            s1_start, s1_end, s2_start, s2_end, s3_start, s3_end)

    # --- Add check summary to Control tab ---
    print("Adding control tab check summary...")
    check_tabs = _add_control_checks(wb, granularity, output_grans)

    # --- Format Control tab ---
    print("Formatting Control tab...")
    format_control_tab(wb['Control'], check_tabs=check_tabs)

    # --- Format Clean Data tabs ---
    for g in _get_available_granularities(granularity, output_grans):
        if g in clean_tabs:
            sn, ly, fd, ld = clean_tabs[g]
            print(f"Formatting {sn}...")
            format_clean_data_tab(wb[sn], ly, fd, ld, g)

    # --- Reorder sheets ---
    _reorder_sheets(wb, granularity)

    # --- Apply formula auditing colors (runs last to override individual cell colors) ---
    print("Applying formula color-coding...")
    apply_formula_coloring(wb)

    # --- Save ---
    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print("Done!")


def _add_control_checks(wb, granularity, output_grans=None):
    """
    Add a Check Summary section to the Control tab (rows 10+).

    Each analysis tab stores its own aggregate check in cell B1.
    This function pulls those values into the Control tab so a user can
    see at a glance whether the whole workbook reconciles to 0.

    Returns the list of (label, tab_name) tuples written, so the formatter
    knows how many rows were added.
    """
    ws = wb['Control']

    # Determine which tabs were actually created (order: retention then cohort)
    check_tabs = []
    for g in _get_available_granularities(granularity, output_grans):
        ret_name = f"{g.capitalize()} Retention"
        if ret_name in wb.sheetnames:
            check_tabs.append((f"{g.capitalize()} Retention Check", ret_name))
    for g in ['annual', 'quarterly']:
        coh_name = f"{g.capitalize()} Cohort"
        if coh_name in wb.sheetnames:
            check_tabs.append((f"{g.capitalize()} Cohort Check", coh_name))

    R_HDR = 10
    ws.cell(row=R_HDR, column=2, value="Check Summary")
    ws.cell(row=R_HDR, column=3, value="Value (= 0)")

    check_cell_refs = []
    for i, (label, tab_name) in enumerate(check_tabs):
        r = R_HDR + 1 + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=f"='{tab_name}'!B1")
        check_cell_refs.append(f"C{r}")

    r_total = R_HDR + 1 + len(check_tabs)
    ws.cell(row=r_total, column=2, value="Total Check (= 0)")
    if check_cell_refs:
        ws.cell(row=r_total, column=3,
                value=f"=SUM({','.join(check_cell_refs)})")

    return check_tabs


def _create_control_tab(wb, config):
    """Create the Control configuration tab."""
    ws = wb.active
    ws.title = "Control"

    ws.cell(row=3, column=2, value="Raw Data Type:")
    ws.cell(row=3, column=3, value="ARR")

    ws.cell(row=4, column=2, value="Scale Factor (Divide By):")
    ws.cell(row=4, column=3, value=config['scale_factor'])

    ws.cell(row=5, column=2, value="Raw Data Timing Interval:")
    ws.cell(row=5, column=3, value=config['time_granularity'].capitalize())

    # Map month number to name
    month_names = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }
    ws.cell(row=6, column=2, value="Fiscal Year End:")
    ws.cell(row=6, column=3, value=month_names[config['fiscal_year_end_month']])

    ws.cell(row=7, column=2, value="Fiscal Year Month #:")
    ws.cell(row=7, column=3, value=config['fiscal_year_end_month'])


def _copy_raw_data(wb, src_wb, config):
    """Copy the raw data sheet from the source workbook."""
    src_ws = src_wb[config['raw_data_sheet']]

    # Add a separator tab
    ws_sep = wb.create_sheet("Raw Data>>")
    ws_sep.cell(row=1, column=1, value="Raw Data >>")

    # Copy the data
    ws = wb.create_sheet(config['raw_data_sheet'])
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)


def _extract_uniques(ws, config):
    """Extract sorted unique dates and customers from raw data."""
    date_col_idx = openpyxl.utils.column_index_from_string(config['date_col'])
    cust_col_idx = openpyxl.utils.column_index_from_string(config['customer_id_col'])
    first_row = config['raw_data_first_row']

    dates = set()
    customers = set()

    for row in ws.iter_rows(min_row=first_row, values_only=False):
        date_val = row[date_col_idx - 1].value
        cust_val = row[cust_col_idx - 1].value
        if date_val is not None:
            dates.add(date_val)
        if cust_val is not None:
            customers.add(cust_val)

    return sorted(dates), sorted(customers)


def _compute_quarterly_dates(monthly_dates, config):
    """
    Compute quarterly period dates from monthly dates.
    Returns dates that are at fiscal quarter boundaries.
    """
    fy_month = config['fiscal_year_end_month']
    quarterly = []
    for d in monthly_dates:
        import datetime
        if isinstance(d, datetime.datetime):
            d_date = d.date()
        else:
            d_date = d
        month = d_date.month
        # Check if this is a quarter-end month
        if (fy_month - month) % 3 == 0:
            quarterly.append(d)
    return quarterly


def _compute_annual_dates(monthly_dates, config):
    """Compute annual period dates from monthly dates."""
    fy_month = config['fiscal_year_end_month']
    annual = []
    for d in monthly_dates:
        import datetime
        if isinstance(d, datetime.datetime):
            d_date = d.date()
        else:
            d_date = d
        if d_date.month == fy_month:
            annual.append(d)
    return annual


def _compute_annual_from_quarterly_dates(quarterly_dates, config):
    """Compute annual dates from quarterly dates (Q4 = fiscal year end)."""
    fy_month = config['fiscal_year_end_month']
    annual = []
    for d in quarterly_dates:
        import datetime
        if isinstance(d, datetime.datetime):
            d_date = d.date()
        else:
            d_date = d
        if d_date.month == fy_month:
            annual.append(d)
    return annual


def _build_filter_blocks(config):
    """
    Build the filter block definitions for retention and cohort tabs.

    Returns a list of dicts with title and filter values.
    The first block is always "Total Business" with all wildcards.
    Additional blocks use filter_breakouts from config.
    """
    attrs = config['attributes']
    attr_names = list(attrs.keys())

    # Block 1: Total Business (all wildcards)
    total_block = {
        'title': 'Total Business',
        'filters': {name: '<>' for name in attr_names}
    }
    total_block['filters']['Cohort'] = '<>'

    blocks = [total_block]

    # Additional breakout blocks from config
    for breakout in config.get('filter_breakouts', []):
        block_filters = {name: '<>' for name in attr_names}
        block_filters['Cohort'] = '<>'
        block_filters.update(breakout.get('filters', {}))
        blocks.append({
            'title': breakout.get('title', 'Filtered'),
            'filters': block_filters
        })

    return blocks


def _get_available_granularities(base_granularity, output_granularities=None):
    """
    Get the granularities to output, in order.

    If output_granularities is provided (from the web wizard), use that list
    filtered to only those that are actually possible given the base data.
    Otherwise fall back to the original behaviour: all granularities from
    the base up to annual.
    """
    all_from_base = {
        'monthly': ['monthly', 'quarterly', 'annual'],
        'quarterly': ['quarterly', 'annual'],
        'annual': ['annual'],
    }[base_granularity]

    if output_granularities:
        # Keep only those the user asked for AND that the data supports
        return [g for g in all_from_base if g in output_granularities]

    return all_from_base


def _reorder_sheets(wb, granularity):
    """Reorder sheets to match the expected tab order."""
    desired_order = ['Control']

    if granularity == 'monthly':
        desired_order.extend([
            'Annual Top Customer Analysis',
            'Annual Cohort', 'Quarterly Cohort',
            'Annual Retention', 'Quarterly Retention', 'Monthly Retention',
            'Clean Annual Data', 'Clean Quarterly Data', 'Clean Monthly Data',
        ])
    elif granularity == 'quarterly':
        desired_order.extend([
            'Annual Top Customer Analysis',
            'Annual Cohort', 'Quarterly Cohort',
            'Annual Retention', 'Quarterly Retention',
            'Clean Annual Data', 'Clean Quarterly Data',
        ])
    else:
        desired_order.extend([
            'Annual Top Customer Analysis',
            'Annual Cohort',
            'Annual Retention',
            'Clean Annual Data',
        ])

    desired_order.extend(['Raw Data>>', wb.sheetnames[-1]])  # raw data sheet last

    # Build the final ordered list: desired first, then any extras
    final_order = []
    for name in desired_order:
        if name in wb.sheetnames and name not in final_order:
            final_order.append(name)
    for name in wb.sheetnames:
        if name not in final_order:
            final_order.append(name)

    # Directly set the sheet order via openpyxl's internal list
    wb._sheets = [wb[name] for name in final_order]
