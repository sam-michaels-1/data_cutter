"""
Formatting module for all Data Pack tabs.

Applies fonts, number formats, alignment, bold, fill colors,
freeze panes, and column widths to match the reference file.
"""
import re
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from .utils import col_letter


# ---- Shared styles ----
BASE_FONT = Font(name='Times New Roman', size=10)
BOLD_FONT = Font(name='Times New Roman', size=10, bold=True)
BLUE_FONT = Font(name='Times New Roman', size=10, color='0000FF')
TITLE_FONT = Font(name='Times New Roman', size=10, bold=True)

YELLOW_FILL = PatternFill(start_color='FFFFFFCC', end_color='FFFFFFCC', fill_type='solid')
GREEN_CF_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_CF_FILL   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# ---- Formula auditing colors ----
# Green  = formula referencing another sheet (contains '!')
# Purple = direct same-sheet cell link only (=CellRef)
# Blue   = hardcoded numeric value (no '=')
_GREEN_COLOR  = '00B050'
_PURPLE_COLOR = '7030A0'
_BLUE_COLOR   = '0000FF'

# Matches a formula that is ONLY a single cell reference: =A1, =$B$3, =AA12, etc.
_DIRECT_LINK_RE = re.compile(r'^=\$?[A-Z]+\$?\d+$', re.IGNORECASE)

CENTER = Alignment(horizontal='center')
CENTER_CONT = Alignment(horizontal='centerContinuous')
LEFT = Alignment(horizontal='left')
RIGHT = Alignment(horizontal='right')

# Number formats matching the reference file
NF_DOLLAR = '* _(* "$"\\ #,##0_);_(* "$"\\ \\(#,##0\\);* \\-_);* @_)'
NF_DOLLAR_DEC = '* _(* "$"\\ #,##0.0_);_(* "$"\\ \\(#,##0.0\\);* \\-_);* @_)'
NF_NUMBER = '* #,##0_);* \\(#,##0\\);* \\-_);* @_)'
NF_NUMBER_DEC = '* #,##0.0_);* \\(#,##0.0\\);* \\-_);* @_)'
NF_PCT = '* #,##0%_);* \\(#,##0%\\);* \\-_)'
NF_PCT_DEC = '* #,##0.0%_);* \\(#,##0.0%\\);* \\-_)'
NF_TIMES = '* #,##0.0\\x_);* \\(#,##0.0\\x\\);* \\-_);* @_)'
NF_DATE = "mmm\\ \\'yy"


def _set_font_all(ws, max_row, max_col, font=None):
    """Set font for all cells in range."""
    if font is None:
        font = BASE_FONT
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = Font(name=font.name, size=font.size,
                             bold=cell.font.bold or False,
                             italic=cell.font.italic or False,
                             color=cell.font.color)


def format_control_tab(ws, check_tabs=None):
    """
    Format the Control tab.

    check_tabs: list of (label, tab_name) tuples written by _add_control_checks(),
                used to format the check summary section starting at row 10.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BASE_FONT

    # Bold labels in column B (config section rows 3-7)
    for r in range(3, 8):
        ws.cell(row=r, column=2).font = BOLD_FONT

    # Blue font + yellow fill for input cells in column C (config section)
    for r in range(3, 8):
        cell = ws.cell(row=r, column=3)
        cell.font = BLUE_FONT
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER

    # Column widths
    ws.column_dimensions['A'].width = 8.73
    ws.column_dimensions['B'].width = 25.63
    ws.column_dimensions['C'].width = 15.63

    # ---- Check Summary section (rows 10+) ----
    if check_tabs:
        R_HDR = 10
        # Header row
        ws.cell(row=R_HDR, column=2).font = BOLD_FONT
        ws.cell(row=R_HDR, column=3).font = BOLD_FONT
        ws.cell(row=R_HDR, column=3).alignment = CENTER

        # One row per analysis tab
        for i in range(len(check_tabs)):
            r = R_HDR + 1 + i
            ws.cell(row=r, column=2).font = BASE_FONT
            ws.cell(row=r, column=3).number_format = NF_NUMBER
            ws.cell(row=r, column=3).alignment = CENTER

        # Total row
        r_total = R_HDR + 1 + len(check_tabs)
        ws.cell(row=r_total, column=2).font = BOLD_FONT
        ws.cell(row=r_total, column=3).font = BOLD_FONT
        ws.cell(row=r_total, column=3).number_format = NF_NUMBER
        ws.cell(row=r_total, column=3).alignment = CENTER

        # Conditional format on total: green = 0, red ≠ 0
        total_ref = f"C{r_total}"
        ws.conditional_formatting.add(
            total_ref, CellIsRule(operator='equal', formula=['0'], fill=GREEN_CF_FILL))
        ws.conditional_formatting.add(
            total_ref, CellIsRule(operator='notEqual', formula=['0'], fill=RED_CF_FILL))


def format_clean_data_tab(ws, layout, first_data_row, last_data_row, granularity):
    """Format a clean data tab."""
    max_col = layout['new_biz_end']

    # Set base font for all cells
    for row in ws.iter_rows(min_row=1, max_row=last_data_row, max_col=max_col):
        for cell in row:
            cell.font = Font(name='Times New Roman', size=10,
                             bold=cell.font.bold if cell.font.bold else False)

    # Row 1: totals - bold
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).font = BOLD_FONT

    # Row 5: section headers - bold
    for c in range(1, max_col + 1):
        cell = ws.cell(row=5, column=c)
        if cell.value:
            cell.font = BOLD_FONT

    # Row 6: column headers - bold, center
    for c in range(1, max_col + 1):
        cell = ws.cell(row=6, column=c)
        if cell.value:
            cell.font = BOLD_FONT
            cell.alignment = CENTER

    # Date columns in row 6 - date format
    for c in range(layout['arr_start'], layout['arr_end'] + 1):
        ws.cell(row=6, column=c).number_format = NF_DATE

    # ARR data - number format
    for r in range(first_data_row, last_data_row + 1):
        for c in range(layout['arr_start'], layout['arr_end'] + 1):
            ws.cell(row=r, column=c).number_format = NF_NUMBER

    # Derived sections - number format
    for section in ['churn', 'downsell', 'upsell', 'new_biz']:
        s = layout[f'{section}_start']
        e = layout[f'{section}_end']
        for r in range(first_data_row, last_data_row + 1):
            for c in range(s, e + 1):
                ws.cell(row=r, column=c).number_format = NF_NUMBER

    # Row 1 totals - number format
    for c in range(layout['arr_start'], max_col + 1):
        ws.cell(row=1, column=c).number_format = NF_NUMBER

    # Freeze panes at the first data cell
    ws.freeze_panes = f"{col_letter(layout['arr_start'])}{first_data_row}"


def format_retention_tab(ws, config, filter_blocks, num_derived, num_attrs,
                         s1_label, s1_start, s1_end,
                         s2_label, s2_start, s2_end,
                         s3_label, s3_start, s3_end,
                         filter_start, cohort_fc):
    """Format a retention tab."""
    max_col = s3_end
    max_row = 5 + len(filter_blocks) * 19

    # Set base font
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = Font(name='Times New Roman', size=10,
                             bold=cell.font.bold if cell.font.bold else False)

    # Freeze panes at first data cell of first block
    ws.freeze_panes = f"{col_letter(s1_start)}8"

    for block_idx in range(len(filter_blocks)):
        start = 5 + block_idx * 19

        r_title     = start + 0
        r_sections  = start + 1
        r_header    = start + 2
        r_bop       = start + 3
        r_churn     = start + 4
        r_downsell  = start + 5
        r_upsell    = start + 6
        r_retained  = start + 7
        r_new_logo  = start + 8
        r_eop       = start + 9
        r_growth    = start + 10
        r_check     = start + 11
        r_lost_ret  = start + 13
        r_punit_ret = start + 14
        r_net_ret   = start + 15
        r_nl_pct    = start + 16
        r_nl_growth = start + 17

        # Title - bold
        ws.cell(row=r_title, column=s1_label).font = BOLD_FONT

        # Section headers - bold, centerContinuous
        for col in [s1_label, s2_label, s3_label]:
            cell = ws.cell(row=r_sections, column=col)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_CONT

        # Column headers - bold, center
        for attr_idx in range(num_attrs):
            ws.cell(row=r_header, column=filter_start + attr_idx).font = BOLD_FONT
            ws.cell(row=r_header, column=filter_start + attr_idx).alignment = CENTER
        ws.cell(row=r_header, column=cohort_fc).font = BOLD_FONT
        ws.cell(row=r_header, column=cohort_fc).alignment = CENTER

        # Date headers - bold, center, date format
        for i in range(num_derived):
            for s_start in [s1_start, s2_start, s3_start]:
                cell = ws.cell(row=r_header, column=s_start + i)
                cell.font = BOLD_FONT
                cell.alignment = CENTER
                cell.number_format = NF_DATE

        # Filter values - bold on BoP row, center
        for attr_idx in range(num_attrs):
            fc = filter_start + attr_idx
            ws.cell(row=r_bop, column=fc).font = BOLD_FONT
            ws.cell(row=r_bop, column=fc).alignment = CENTER
            for r in range(r_churn, r_nl_growth + 1):
                ws.cell(row=r, column=fc).alignment = CENTER
        ws.cell(row=r_bop, column=cohort_fc).font = BOLD_FONT
        ws.cell(row=r_bop, column=cohort_fc).alignment = CENTER

        # --- Section 1 (Net Retention) formatting ---
        # Bold rows: BoP, Retained, EoP, Net Retention
        for r in [r_bop, r_retained, r_eop, r_net_ret]:
            ws.cell(row=r, column=s1_label).font = BOLD_FONT
            for c in range(s1_start, s1_end + 1):
                ws.cell(row=r, column=c).font = BOLD_FONT

        # Left-aligned label rows
        for r in [r_churn, r_downsell, r_upsell, r_new_logo, r_growth, r_check,
                  r_net_ret]:
            ws.cell(row=r, column=s1_label).alignment = LEFT

        # Number formats - Section 1
        for i in range(num_derived):
            c = s1_start + i
            # Dollar format for ARR rows
            for r in [r_bop, r_retained, r_eop]:
                ws.cell(row=r, column=c).number_format = NF_DOLLAR
            # Number format for component rows
            for r in [r_churn, r_downsell, r_upsell, r_new_logo, r_check]:
                ws.cell(row=r, column=c).number_format = NF_NUMBER
            # Percentage format
            for r in [r_growth]:
                ws.cell(row=r, column=c).number_format = NF_PCT_DEC
            for r in [r_lost_ret, r_punit_ret, r_net_ret, r_nl_pct]:
                ws.cell(row=r, column=c).number_format = NF_PCT_DEC

        # --- Section 2 (Customer Retention) formatting ---
        # Bold rows
        for r in [r_bop, r_downsell, r_retained]:  # BoP, Retained, EoP
            ws.cell(row=r, column=s2_label).font = BOLD_FONT
            for c in range(s2_start, s2_end + 1):
                ws.cell(row=r, column=c).font = BOLD_FONT

        # Left-aligned labels
        for r in [r_churn, r_upsell, r_new_logo, r_check]:
            ws.cell(row=r, column=s2_label).alignment = LEFT

        # Number format - Section 2
        for i in range(num_derived):
            c = s2_start + i
            for r in [r_bop, r_churn, r_downsell, r_upsell, r_retained, r_check]:
                ws.cell(row=r, column=c).number_format = NF_NUMBER
            ws.cell(row=r_new_logo, column=c).number_format = NF_PCT
            ws.cell(row=r_lost_ret, column=c).number_format = NF_PCT
            ws.cell(row=r_punit_ret, column=c).number_format = NF_PCT

        # --- Section 3 (ARR/Customer) formatting ---
        # Bold rows
        for r in [r_bop, r_upsell, r_new_logo]:  # BoP, Retained, EoP
            ws.cell(row=r, column=s3_label).font = BOLD_FONT
            for c in range(s3_start, s3_end + 1):
                ws.cell(row=r, column=c).font = BOLD_FONT

        # Left-aligned labels
        for r in [r_churn, r_downsell, r_retained, r_eop]:
            ws.cell(row=r, column=s3_label).alignment = LEFT

        # Number format - Section 3
        for i in range(num_derived):
            c = s3_start + i
            # Dollar decimal for ARR/Customer
            for r in [r_bop, r_upsell, r_new_logo]:  # BoP, Retained, EoP
                ws.cell(row=r, column=c).number_format = NF_DOLLAR_DEC
            for r in [r_churn, r_downsell, r_retained]:
                ws.cell(row=r, column=c).number_format = NF_NUMBER_DEC
            ws.cell(row=r_eop, column=c).number_format = NF_PCT
            # Times format for New Logo vs Churn, New Logo vs Retained
            ws.cell(row=r_lost_ret, column=c).number_format = NF_TIMES
            ws.cell(row=r_punit_ret, column=c).number_format = NF_TIMES

            # Right-align all S3 data
            for r in range(r_bop, r_nl_growth + 1):
                ws.cell(row=r, column=c).alignment = RIGHT


def format_cohort_tab(ws, config, filter_blocks, num_dates, num_cohorts,
                      num_attrs,
                      q_col, y_col, filter_start, cohort_label_col,
                      s1_start, s1_end, s2_start, s2_end,
                      s3_label, s3_start_val, s3_data_start, s3_data_end,
                      s4_label, s4_start_val, s4_data_start, s4_data_end,
                      granularity):
    """Format a cohort tab."""
    max_col = s4_data_end

    for block_idx in range(len(filter_blocks)):
        block_start = 6 + block_idx * (num_cohorts + 9)
        r_section_headers = block_start + 2
        r_headers = block_start + 3
        first_cohort_row = r_headers + 1
        last_cohort_row = first_cohort_row + num_cohorts - 1
        r_total = last_cohort_row + 1
        r_median = r_total + 1
        r_weighted = r_median + 1
        r_check = r_weighted + 1

        # Title - bold
        ws.cell(row=block_start, column=q_col).font = BOLD_FONT

        # Section headers - bold, centerContinuous
        for col in [s1_start, s2_start, s3_label, s4_label]:
            cell = ws.cell(row=r_section_headers, column=col)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_CONT

        # Column headers - bold, center
        for col in [q_col, y_col]:
            ws.cell(row=r_headers, column=col).font = BOLD_FONT
            ws.cell(row=r_headers, column=col).alignment = CENTER
        for attr_idx in range(num_attrs):
            ws.cell(row=r_headers, column=filter_start + attr_idx).font = BOLD_FONT
            ws.cell(row=r_headers, column=filter_start + attr_idx).alignment = CENTER
        ws.cell(row=r_headers, column=cohort_label_col).font = BOLD_FONT
        ws.cell(row=r_headers, column=cohort_label_col).alignment = CENTER

        # ARR/Customer period headers - center
        for i in range(num_dates):
            ws.cell(row=r_headers, column=s1_start + i).alignment = CENTER
            ws.cell(row=r_headers, column=s1_start + i).font = BOLD_FONT
            ws.cell(row=r_headers, column=s2_start + i).alignment = CENTER
            ws.cell(row=r_headers, column=s2_start + i).font = BOLD_FONT

        # Retention headers - center
        ws.cell(row=r_headers, column=s3_start_val).alignment = CENTER
        ws.cell(row=r_headers, column=s3_start_val).font = BOLD_FONT
        ws.cell(row=r_headers, column=s4_start_val).alignment = CENTER
        ws.cell(row=r_headers, column=s4_start_val).font = BOLD_FONT
        for i in range(num_dates):
            ws.cell(row=r_headers, column=s3_data_start + i).alignment = CENTER
            ws.cell(row=r_headers, column=s3_data_start + i).font = BOLD_FONT
            ws.cell(row=r_headers, column=s4_data_start + i).alignment = CENTER
            ws.cell(row=r_headers, column=s4_data_start + i).font = BOLD_FONT

        # Data rows formatting
        for r in range(first_cohort_row, r_check + 1):
            # Quarter/Year - center
            ws.cell(row=r, column=q_col).alignment = CENTER
            ws.cell(row=r, column=y_col).alignment = CENTER
            # Filters - center
            for attr_idx in range(num_attrs):
                ws.cell(row=r, column=filter_start + attr_idx).alignment = CENTER
            # Cohort label - center
            ws.cell(row=r, column=cohort_label_col).alignment = CENTER

            # ARR data - number format
            for i in range(num_dates):
                ws.cell(row=r, column=s1_start + i).number_format = NF_NUMBER
                ws.cell(row=r, column=s2_start + i).number_format = NF_NUMBER

            # Starting value - dollar format for ARR, number for customers
            ws.cell(row=r, column=s3_start_val).number_format = NF_DOLLAR
            ws.cell(row=r, column=s4_start_val).number_format = NF_NUMBER

            # Retention percentages
            for i in range(num_dates):
                ws.cell(row=r, column=s3_data_start + i).number_format = NF_PCT
                ws.cell(row=r, column=s4_data_start + i).number_format = NF_PCT

        # Bold summary rows
        for r in [r_total]:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.font = BOLD_FONT

        # Set base font for all
        for row in ws.iter_rows(min_row=1, max_row=r_check, max_col=max_col):
            for cell in row:
                if cell.font.name != 'Times New Roman':
                    cell.font = Font(name='Times New Roman', size=10,
                                     bold=cell.font.bold if cell.font.bold else False)


def format_top_customers_tab(ws, config, clean_layout,
                             first_customer_row, last_customer_row,
                             r_top_total, r_other, r_total, r_memo_start,
                             num_dates,
                             rank_num_col, cust_id_col, attr_start, num_attrs,
                             cohort_col,
                             s1_start, s1_end, s2_start, s2_end, s3_start, s3_end):
    """Format the Annual Top Customer Analysis tab."""
    max_col = s3_end

    # Set base font for all
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.font = Font(name='Times New Roman', size=10,
                             bold=cell.font.bold if cell.font.bold else False)

    # Row 5: Section headers - bold, centerContinuous
    ws.cell(row=5, column=s1_start).font = BOLD_FONT
    ws.cell(row=5, column=s1_start).alignment = CENTER_CONT
    ws.cell(row=5, column=s2_start).font = BOLD_FONT
    ws.cell(row=5, column=s2_start).alignment = CENTER_CONT
    ws.cell(row=5, column=s3_start).font = BOLD_FONT
    ws.cell(row=5, column=s3_start).alignment = CENTER_CONT

    # Row 6: Column headers - bold, center
    for c in [cust_id_col] + list(range(attr_start, attr_start + num_attrs)) + [cohort_col]:
        ws.cell(row=6, column=c).font = BOLD_FONT
        ws.cell(row=6, column=c).alignment = CENTER

    for i in range(num_dates):
        ws.cell(row=6, column=s1_start + i).font = BOLD_FONT
        ws.cell(row=6, column=s1_start + i).alignment = CENTER
    for i in range(num_dates - 1):
        ws.cell(row=6, column=s2_start + i).font = BOLD_FONT
        ws.cell(row=6, column=s2_start + i).alignment = CENTER
    for i in range(num_dates):
        ws.cell(row=6, column=s3_start + i).font = BOLD_FONT
        ws.cell(row=6, column=s3_start + i).alignment = CENTER

    # Data rows
    for r in range(first_customer_row, last_customer_row + 1):
        ws.cell(row=r, column=rank_num_col).alignment = CENTER
        for attr_idx in range(num_attrs):
            ws.cell(row=r, column=attr_start + attr_idx).alignment = CENTER
        ws.cell(row=r, column=cohort_col).alignment = CENTER

        # Number formats
        for i in range(num_dates):
            ws.cell(row=r, column=s1_start + i).number_format = NF_NUMBER
        for i in range(num_dates - 1):
            ws.cell(row=r, column=s2_start + i).number_format = NF_PCT
            ws.cell(row=r, column=s2_start + i).alignment = RIGHT
        for i in range(num_dates):
            ws.cell(row=r, column=s3_start + i).number_format = NF_PCT_DEC
            ws.cell(row=r, column=s3_start + i).alignment = RIGHT

    # Summary rows - bold
    for r in [r_top_total, r_total]:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.font = BOLD_FONT
        # Number formats for summary
        for i in range(num_dates):
            ws.cell(row=r, column=s1_start + i).number_format = NF_NUMBER
        for i in range(num_dates - 1):
            ws.cell(row=r, column=s2_start + i).number_format = NF_PCT
        for i in range(num_dates):
            ws.cell(row=r, column=s3_start + i).number_format = NF_PCT_DEC

    # Other row (not bold)
    for i in range(num_dates):
        ws.cell(row=r_other, column=s1_start + i).number_format = NF_NUMBER
    for i in range(num_dates - 1):
        ws.cell(row=r_other, column=s2_start + i).number_format = NF_PCT
    for i in range(num_dates):
        ws.cell(row=r_other, column=s3_start + i).number_format = NF_PCT_DEC

    # Concentration memo rows
    for tier_idx in range(4):
        r = r_memo_start + 1 + tier_idx
        for i in range(num_dates):
            ws.cell(row=r, column=s1_start + i).number_format = NF_NUMBER
        for i in range(num_dates - 1):
            ws.cell(row=r, column=s2_start + i).number_format = NF_PCT
        for i in range(num_dates):
            ws.cell(row=r, column=s3_start + i).number_format = NF_PCT_DEC


# ---------------------------------------------------------------------------
# Formula auditing color-coding (runs last, after all other formatting)
# ---------------------------------------------------------------------------

def apply_formula_coloring(wb, skip_sheets=None):
    """
    Walk every cell in the workbook and apply formula-auditing colors:

      Green  (#00B050) — formula that references another sheet (contains '!')
      Purple (#7030A0) — formula that is ONLY a direct cell link, e.g. =B9
      Blue   (#0000FF) — hardcoded numeric value (int or float, not a formula)
      (unchanged)      — complex same-sheet formula, or any string/text label

    All other font properties (name, size, bold, italic) are preserved.
    skip_sheets: optional list of sheet titles to leave untouched.
    """
    skip = set(skip_sheets or [])

    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # Preserve existing font attributes
                old   = cell.font
                fname = old.name  or 'Times New Roman'
                fsize = old.size  or 10
                bold  = old.bold  or False
                italic = old.italic or False

                val = str(cell.value)

                if val.startswith('='):
                    body = val[1:]
                    if '!' in body:
                        # Cross-sheet reference → green
                        cell.font = Font(name=fname, size=fsize, bold=bold,
                                         italic=italic, color=_GREEN_COLOR)
                    elif _DIRECT_LINK_RE.match(val):
                        # Pure same-sheet cell link → purple
                        cell.font = Font(name=fname, size=fsize, bold=bold,
                                         italic=italic, color=_PURPLE_COLOR)
                    # else: complex same-sheet formula → leave color unchanged

                elif isinstance(cell.value, (int, float)):
                    # Hardcoded numeric literal → blue
                    # (datetime objects are NOT int/float, so date headers stay black)
                    cell.font = Font(name=fname, size=fsize, bold=bold,
                                     italic=italic, color=_BLUE_COLOR)

                # String literals (labels, IDs, text) → leave unchanged
