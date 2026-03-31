"""
Top Customer Analysis tab generator.

Shows the top N customers ranked by latest-period ARR.

Three sections:
  1. ARR over time
  2. % YoY Growth
  3. % of Total

Rows:
  - Top N individual customers (looked up by rank)
  - Top N Total
  - Other Customers
  - Total ARR
  - Concentration memo (Top 3, 5, 10, 25)
"""
from openpyxl.utils import get_column_letter
from .utils import col_letter


TOP_N = 25
CONCENTRATION_TIERS = [3, 5, 10, 25]


def generate_top_customers_tab(wb, config, clean_sheet_name, clean_layout,
                               first_data_row, last_data_row):
    """Generate the Annual Top Customer Analysis tab."""
    sheet_name = "Annual Top Customer Analysis"
    ws = wb.create_sheet(sheet_name)

    num_dates = clean_layout['num_dates']
    num_attrs = clean_layout['num_attrs']

    # Column layout
    # B: rank number, C: Customer ID, D-F: attributes, G: cohort
    rank_num_col = 2  # B
    cust_id_col = 3   # C
    attr_start = 4    # D
    attr_end = attr_start + num_attrs - 1
    cohort_col = attr_end + 1

    # Section 1: ARR
    s1_start = cohort_col + 1
    s1_end = s1_start + num_dates - 1

    # Section 2: % YoY Growth (one fewer column than ARR)
    s2_start = s1_end + 2  # gap
    s2_end = s2_start + num_dates - 2  # n-1 growth periods

    # Section 3: % of Total
    s3_start = s2_end + 2  # gap
    s3_end = s3_start + num_dates - 1

    # Clean data references
    cdr_first = first_data_row
    cdr_last = last_data_row
    clean_rank_col = col_letter(clean_layout['rank'])
    clean_cust_id_col = col_letter(clean_layout['cust_id'])

    # --- Row 3: Units (A3=label, B3=formula) ---
    ws.cell(row=3, column=1, value="Units")
    ws.cell(row=3, column=rank_num_col, value="=Control!$C$4")
    units_cell = f"${col_letter(rank_num_col)}$3"

    # --- Row 5: Section headers ---
    metric_label = "ARR" if config.get("data_type", "arr") == "arr" else "Revenue"
    ws.cell(row=5, column=s1_start, value=metric_label)
    ws.cell(row=5, column=s2_start, value="% YoY Growth")
    ws.cell(row=5, column=s3_start, value="% of Total")

    # --- Row 6: Column headers ---
    ws.cell(row=6, column=cust_id_col, value="Customer ID")
    attr_names = list(config['attributes'].keys())
    for i, name in enumerate(attr_names):
        ws.cell(row=6, column=attr_start + i, value=name)
    ws.cell(row=6, column=cohort_col, value="Annual Cohort")

    # Period headers for section 1 (reference clean data)
    for i in range(num_dates):
        clean_col = clean_layout['arr_start'] + i
        ws.cell(row=6, column=s1_start + i,
                value=f"='{clean_sheet_name}'!{col_letter(clean_col)}6")

    # Period headers for section 2 (reference section 1, offset by 1)
    for i in range(num_dates - 1):
        ws.cell(row=6, column=s2_start + i,
                value=f"={col_letter(s1_start + i + 1)}6")

    # Period headers for section 3 (reference section 1)
    for i in range(num_dates):
        ws.cell(row=6, column=s3_start + i,
                value=f"={col_letter(s1_start + i)}6")

    # --- Data rows (Top N customers) ---
    first_customer_row = 7
    last_customer_row = first_customer_row + TOP_N - 1

    for rank in range(1, TOP_N + 1):
        row = first_customer_row + rank - 1

        # Rank number
        if rank == 1:
            ws.cell(row=row, column=rank_num_col, value=1)
        else:
            ws.cell(row=row, column=rank_num_col,
                    value=f"={col_letter(rank_num_col)}{row-1}+1")

        # Customer ID via XLOOKUP by rank
        rnl = col_letter(rank_num_col)
        ws.cell(row=row, column=cust_id_col,
                value=f"=_xlfn.XLOOKUP(${rnl}{row},"
                      f"'{clean_sheet_name}'!${clean_rank_col}${cdr_first}:${clean_rank_col}${cdr_last},"
                      f"'{clean_sheet_name}'!{col_letter(clean_layout['cust_id'])}${cdr_first}"
                      f":{col_letter(clean_layout['cust_id'])}${cdr_last})")

        # Attributes via XLOOKUP by rank
        for attr_idx in range(num_attrs):
            clean_attr_col = clean_layout['attr_start'] + attr_idx
            ws.cell(row=row, column=attr_start + attr_idx,
                    value=f"=_xlfn.XLOOKUP(${rnl}{row},"
                          f"'{clean_sheet_name}'!${clean_rank_col}${cdr_first}:${clean_rank_col}${cdr_last},"
                          f"'{clean_sheet_name}'!{col_letter(clean_attr_col)}${cdr_first}"
                          f":{col_letter(clean_attr_col)}${cdr_last})")

        # Cohort via XLOOKUP by rank
        ws.cell(row=row, column=cohort_col,
                value=f"=_xlfn.XLOOKUP(${rnl}{row},"
                      f"'{clean_sheet_name}'!${clean_rank_col}${cdr_first}:${clean_rank_col}${cdr_last},"
                      f"'{clean_sheet_name}'!{col_letter(clean_layout['cohort'])}${cdr_first}"
                      f":{col_letter(clean_layout['cohort'])}${cdr_last})")

        # Section 1: ARR per period (SUMIFS by customer ID)
        cidl = col_letter(cust_id_col)
        for i in range(num_dates):
            clean_col = clean_layout['arr_start'] + i
            ws.cell(row=row, column=s1_start + i,
                    value=f"=SUMIFS('{clean_sheet_name}'!{col_letter(clean_col)}${cdr_first}"
                          f":{col_letter(clean_col)}${cdr_last},"
                          f"'{clean_sheet_name}'!${clean_cust_id_col}${cdr_first}"
                          f":${clean_cust_id_col}${cdr_last},${cidl}{row})/{units_cell}")

        # Section 2: % YoY Growth
        for i in range(num_dates - 1):
            curr_col = col_letter(s1_start + i + 1)
            prev_col = col_letter(s1_start + i)
            ws.cell(row=row, column=s2_start + i,
                    value=f'=IFERROR({curr_col}{row}/{prev_col}{row}-1,"n.a.")')

        # Section 3: % of Total
        total_row = last_customer_row + 3  # skip Top N Total, Other
        for i in range(num_dates):
            arr_col = col_letter(s1_start + i)
            ws.cell(row=row, column=s3_start + i,
                    value=f"={arr_col}{row}/{arr_col}${total_row}")

    # --- Summary rows ---
    r_top_total = last_customer_row + 1
    r_other = r_top_total + 1
    r_total = r_other + 1

    # Top N Total
    ws.cell(row=r_top_total, column=cust_id_col, value=f"Top {TOP_N} Customers")
    for i in range(num_dates):
        arr_cl = col_letter(s1_start + i)
        ws.cell(row=r_top_total, column=s1_start + i,
                value=f"=SUM({arr_cl}{first_customer_row}:{arr_cl}{last_customer_row})")

        # Growth
        if i > 0:
            curr = col_letter(s1_start + i)
            prev = col_letter(s1_start + i - 1)
            ws.cell(row=r_top_total, column=s2_start + i - 1,
                    value=f'=IFERROR({curr}{r_top_total}/{prev}{r_top_total}-1,"n.a.")')

        # % of Total
        ws.cell(row=r_top_total, column=s3_start + i,
                value=f"={arr_cl}{r_top_total}/{arr_cl}${r_total}")

    # Other Customers = Total - Top N
    ws.cell(row=r_other, column=cust_id_col, value="(+) Other Customers")
    for i in range(num_dates):
        arr_cl = col_letter(s1_start + i)
        ws.cell(row=r_other, column=s1_start + i,
                value=f"={arr_cl}{r_total}-{arr_cl}{r_top_total}")

        if i > 0:
            curr = col_letter(s1_start + i)
            prev = col_letter(s1_start + i - 1)
            ws.cell(row=r_other, column=s2_start + i - 1,
                    value=f'=IFERROR({curr}{r_other}/{prev}{r_other}-1,"n.a.")')

        ws.cell(row=r_other, column=s3_start + i,
                value=f"={arr_cl}{r_other}/{arr_cl}${r_total}")

    # Total ARR (all customers)
    ws.cell(row=r_total, column=cust_id_col, value=f"Total {metric_label}")
    for i in range(num_dates):
        clean_col = clean_layout['arr_start'] + i
        ws.cell(row=r_total, column=s1_start + i,
                value=f'=SUMIFS(\'{clean_sheet_name}\'!{col_letter(clean_col)}${cdr_first}'
                      f':{col_letter(clean_col)}${cdr_last},'
                      f'\'{clean_sheet_name}\'!${clean_cust_id_col}${cdr_first}'
                      f':${clean_cust_id_col}${cdr_last},"<>")/{units_cell}')

        if i > 0:
            curr = col_letter(s1_start + i)
            prev = col_letter(s1_start + i - 1)
            ws.cell(row=r_total, column=s2_start + i - 1,
                    value=f'=IFERROR({curr}{r_total}/{prev}{r_total}-1,"n.a.")')

    # --- Concentration Memo ---
    r_memo_start = r_total + 2
    ws.cell(row=r_memo_start, column=cust_id_col, value="Memo:")

    for tier_idx, tier in enumerate(CONCENTRATION_TIERS):
        row = r_memo_start + 1 + tier_idx
        ws.cell(row=row, column=rank_num_col, value=tier)
        ws.cell(row=row, column=cust_id_col,
                value=f'="Top "&{col_letter(rank_num_col)}{row}&" Customers Concentration"')

        rnl = col_letter(rank_num_col)
        for i in range(num_dates):
            arr_cl = col_letter(s1_start + i)
            ws.cell(row=row, column=s1_start + i,
                    value=f"=SUMIFS({arr_cl}${first_customer_row}:{arr_cl}${last_customer_row},"
                          f"${rnl}${first_customer_row}:${rnl}${last_customer_row},"
                          f'"<="&${rnl}{row})')

            if i > 0:
                curr = col_letter(s1_start + i)
                prev = col_letter(s1_start + i - 1)
                ws.cell(row=row, column=s2_start + i - 1,
                        value=f'=IFERROR({curr}{row}/{prev}{row}-1,"n.a.")')

            ws.cell(row=row, column=s3_start + i,
                    value=f"={arr_cl}{row}/{arr_cl}${r_total}")

    return sheet_name
