"""
Clean Data tab generators.

Generates:
- Base clean data tab (at raw data granularity) with SUMIFS from raw data
- Aggregated clean data tabs (quarterly from monthly, annual from quarterly/monthly)

Each tab structure:
  Row 1: Totals + fiscal month end reference
  Row 2: Quarter helper
  Row 3: Year helper
  Row 4: empty
  Row 5: Section headers
  Row 6: Column headers (dates, customer info labels)
  Row 7+: Customer data
"""
from openpyxl.utils import get_column_letter
from .utils import (
    col_letter, col_num, compute_clean_layout, make_range, make_cell,
    sumifs, get_yoy_offset
)


# ---------------------------------------------------------------------------
# Base clean data tab (SUMIFS from raw data)
# ---------------------------------------------------------------------------

def generate_base_clean_data(wb, config, unique_dates, unique_customers):
    """
    Generate the base clean data tab at the raw data's native granularity.
    All ARR values are SUMIFS formulas referencing the raw data sheet.
    """
    granularity = config['time_granularity']
    sheet_name = f"Clean {granularity.capitalize()} Data"
    ws = wb.create_sheet(sheet_name)

    raw_sheet = config['raw_data_sheet']
    raw_first = config['raw_data_first_row']
    raw_last = config['raw_data_last_row']
    raw_cust_col = config['customer_id_col']
    raw_date_col = config['date_col']
    raw_arr_col = config['arr_col']
    attrs = config['attributes']  # OrderedDict: {display_name: raw_col_letter}
    num_attrs = len(attrs)
    num_dates = len(unique_dates)
    num_customers = len(unique_customers)
    yoy_offset = get_yoy_offset(granularity)
    fy_month = config['fiscal_year_end_month']

    layout = compute_clean_layout(num_attrs, num_dates, yoy_offset)
    first_data_row = 7
    last_data_row = first_data_row + num_customers - 1

    # --- Row 1: Fiscal month end + column totals ---
    ws.cell(row=1, column=1, value="Fiscal Month End:")
    ws.cell(row=1, column=2, value="=Control!C7")

    # Column totals for all data sections
    for section_key in ['arr', 'churn', 'downsell', 'upsell', 'new_biz']:
        start = layout[f'{section_key}_start']
        end = layout[f'{section_key}_end'] if section_key != 'arr' else layout['arr_end']
        for c in range(start, end + 1):
            cl = col_letter(c)
            ws.cell(row=1, column=c,
                    value=f"=SUM({cl}{first_data_row}:{cl}{last_data_row})")

    # --- Row 2: Quarter calculation ---
    ws.cell(row=2, column=layout['label'], value="Quarter")

    if granularity == 'monthly':
        # For monthly data, calculate fiscal quarter for each month
        for i in range(num_dates):
            c = layout['arr_start'] + i
            cl = col_letter(c)
            ws.cell(row=2, column=c,
                    value=f"=IF(MOD($B$1-MONTH({cl}6),3)=0,"
                          f"INT(MOD(MONTH({cl}6)-(Control!$C$7+1),12)/3)+1,0)")
    elif granularity == 'quarterly':
        # For quarterly data, write quarter numbers
        # First quarter is derived from the first date
        import datetime
        first_date = unique_dates[0]
        if isinstance(first_date, datetime.datetime):
            first_date = first_date.date()
        month = first_date.month
        # Calculate fiscal quarter
        fq = int(((month - (fy_month + 1)) % 12) / 3) + 1
        ws.cell(row=2, column=layout['arr_start'], value=fq)
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=2, column=layout['arr_start'] + i,
                    value=f"=IF({prev_cl}2=4,1,{prev_cl}2+1)")
    elif granularity == 'annual':
        # Annual data: all quarters are Q4 (fiscal year end)
        ws.cell(row=2, column=layout['arr_start'], value=4)
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=2, column=layout['arr_start'] + i,
                    value=f"={prev_cl}2")

    # Copy Quarter row to derived sections
    _copy_helper_rows_to_derived(ws, layout, 2)

    # --- Row 3: Year calculation ---
    ws.cell(row=3, column=layout['label'], value="Year")

    if granularity == 'monthly':
        for i in range(num_dates):
            c = layout['arr_start'] + i
            cl = col_letter(c)
            ws.cell(row=3, column=c, value=f"=IF({cl}2>0,YEAR({cl}6),0)")
    elif granularity == 'quarterly':
        import datetime
        first_date = unique_dates[0]
        if isinstance(first_date, datetime.datetime):
            first_date = first_date.date()
        # For quarterly, year adjusts based on fiscal year
        # If the month is after fiscal year end, it's the next fiscal year
        if first_date.month > fy_month:
            first_year = first_date.year + 1
        else:
            first_year = first_date.year
        ws.cell(row=3, column=layout['arr_start'], value=first_year)
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            prev_q_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=3, column=layout['arr_start'] + i,
                    value=f"=IF({prev_q_cl}2=4,{prev_cl}3+1,{prev_cl}3)")
    elif granularity == 'annual':
        import datetime
        first_date = unique_dates[0]
        if isinstance(first_date, datetime.datetime):
            first_date = first_date.date()
        ws.cell(row=3, column=layout['arr_start'], value=first_date.year)
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=3, column=layout['arr_start'] + i,
                    value=f"=IF({prev_cl}2=4,{prev_cl}3+1,{prev_cl}3)")

    # Copy Year row to derived sections
    _copy_helper_rows_to_derived(ws, layout, 3)

    # --- Row 5: Section headers ---
    ws.cell(row=5, column=layout['attr_start'], value="Customer Identifying Information")
    ws.cell(row=5, column=layout['arr_start'],
            value=f"{granularity.capitalize()} ARR by Date")
    ws.cell(row=5, column=layout['churn_start'], value="Churn?")
    ws.cell(row=5, column=layout['downsell_start'], value="Downsell?")
    ws.cell(row=5, column=layout['upsell_start'], value="Upsell?")
    ws.cell(row=5, column=layout['new_biz_start'], value="New Business Dollars?")

    # --- Row 6: Column headers ---
    # Customer info headers
    ws.cell(row=6, column=layout['cust_id'], value="Customer ID")
    attr_names = list(attrs.keys())
    for i, name in enumerate(attr_names):
        ws.cell(row=6, column=layout['attr_start'] + i, value=name)

    cohort_name = f"{granularity.capitalize()} Cohort"
    ws.cell(row=6, column=layout['cohort'], value=cohort_name)
    ws.cell(row=6, column=layout['rank'], value="Customer Rank #")

    # Date headers in ARR section
    for i, date_val in enumerate(unique_dates):
        ws.cell(row=6, column=layout['arr_start'] + i, value=date_val)

    # Derived section date headers (reference the ARR section, offset by yoy_offset)
    for i in range(layout['num_derived']):
        # Churn dates reference ARR dates offset by yoy_offset
        arr_ref_col = layout['arr_start'] + yoy_offset + i
        arr_ref_cl = col_letter(arr_ref_col)

        churn_col = layout['churn_start'] + i
        ws.cell(row=6, column=churn_col, value=f"={arr_ref_cl}6")

        # Downsell refs churn
        downsell_col = layout['downsell_start'] + i
        churn_cl = col_letter(churn_col)
        ws.cell(row=6, column=downsell_col, value=f"={churn_cl}6")

        # Upsell refs downsell
        upsell_col = layout['upsell_start'] + i
        downsell_cl = col_letter(downsell_col)
        ws.cell(row=6, column=upsell_col, value=f"={downsell_cl}6")

        # New Biz refs upsell
        new_biz_col = layout['new_biz_start'] + i
        upsell_cl = col_letter(upsell_col)
        ws.cell(row=6, column=new_biz_col, value=f"={upsell_cl}6")

    # --- Customer data rows (7+) ---
    # Build raw data range references (used in all SUMIFS)
    raw_arr_range = make_range(raw_sheet, raw_arr_col, raw_first, raw_last)
    raw_date_range = make_range(raw_sheet, raw_date_col, raw_first, raw_last)
    raw_cust_range = make_range(raw_sheet, raw_cust_col, raw_first, raw_last)

    arr_start_cl = col_letter(layout['arr_start'])
    arr_end_cl = col_letter(layout['arr_end'])

    for idx, cust_id in enumerate(unique_customers):
        row = first_data_row + idx
        cust_id_cl = col_letter(layout['cust_id'])

        # Column B: Customer ID (value)
        ws.cell(row=row, column=layout['cust_id'], value=cust_id)

        # Columns C-E (etc.): Attribute lookups via XLOOKUP
        for attr_idx, (attr_name, raw_attr_col) in enumerate(attrs.items()):
            attr_col = layout['attr_start'] + attr_idx
            ws.cell(row=row, column=attr_col,
                    value=f"=_xlfn.XLOOKUP(${cust_id_cl}{row},"
                          f"'{raw_sheet}'!${raw_cust_col}${raw_first}:${raw_cust_col}${raw_last},"
                          f"'{raw_sheet}'!${raw_attr_col}${raw_first}:${raw_attr_col}${raw_last})")

        # Column F: Cohort (first non-zero ARR period)
        cohort_cl = col_letter(layout['cohort'])
        ws.cell(row=row, column=layout['cohort'],
                value=f'=IFERROR(INDEX(${arr_start_cl}$6:${arr_end_cl}$6,'
                      f'MATCH(TRUE,INDEX({arr_start_cl}{row}:{arr_end_cl}{row}<>0,0),0)),"n.a.")')

        # Column G: Rank (by last-period ARR)
        ws.cell(row=row, column=layout['rank'],
                value=f"=RANK({arr_end_cl}{row},"
                      f"${arr_end_cl}${first_data_row}:${arr_end_cl}${last_data_row})")

        # ARR columns: SUMIFS from raw data
        for i in range(num_dates):
            c = layout['arr_start'] + i
            cl = col_letter(c)
            ws.cell(row=row, column=c,
                    value=sumifs(
                        raw_arr_range,
                        (raw_date_range, f"{cl}$6"),
                        (raw_cust_range, f"${cust_id_cl}{row}")
                    ))

        # Derived sections: Churn, Downsell, Upsell, New Business
        for i in range(layout['num_derived']):
            # "prior" = ARR at period i (yoy_offset periods ago)
            prior_col = layout['arr_start'] + i
            prior_cl = col_letter(prior_col)
            # "current" = ARR at period i + yoy_offset
            curr_col = layout['arr_start'] + yoy_offset + i
            curr_cl = col_letter(curr_col)

            prior_ref = f"{prior_cl}{row}"
            curr_ref = f"{curr_cl}{row}"

            # Churn: customer went to zero
            ws.cell(row=row, column=layout['churn_start'] + i,
                    value=f"=IF(AND({curr_ref}=0,{prior_ref}>0),-{prior_ref},0)")

            # Downsell: customer decreased but still active
            ws.cell(row=row, column=layout['downsell_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}>0,{curr_ref}<{prior_ref}),"
                          f"{curr_ref}-{prior_ref},0)")

            # Upsell: customer increased and still active
            ws.cell(row=row, column=layout['upsell_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}>0,{curr_ref}>{prior_ref}),"
                          f"{curr_ref}-{prior_ref},0)")

            # New Business: customer appeared (was zero, now positive)
            ws.cell(row=row, column=layout['new_biz_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}=0),{curr_ref},0)")

    return sheet_name, layout, first_data_row, last_data_row


# ---------------------------------------------------------------------------
# Aggregated clean data tabs (quarterly from monthly, annual from monthly)
# ---------------------------------------------------------------------------

def generate_aggregated_clean_data(wb, config, source_sheet, source_layout,
                                   target_granularity, unique_customers,
                                   target_dates):
    """
    Generate an aggregated clean data tab.
    e.g., Clean Quarterly Data from Clean Monthly Data.

    ARR values use SUMIFS across the source tab's row, matching Quarter/Year.
    Churn/Downsell/Upsell/NewBiz are derived from the aggregated ARR values.
    """
    sheet_name = f"Clean {target_granularity.capitalize()} Data"
    ws = wb.create_sheet(sheet_name)

    num_dates = len(target_dates)
    num_attrs = source_layout['num_attrs']
    yoy_offset = get_yoy_offset(target_granularity)
    fy_month = config['fiscal_year_end_month']
    num_customers = len(unique_customers)

    layout = compute_clean_layout(num_attrs, num_dates, yoy_offset)
    first_data_row = 7
    last_data_row = first_data_row + num_customers - 1

    # Source sheet column references (for SUMIFS across the source row)
    src_arr_start_cl = col_letter(source_layout['arr_start'])
    src_arr_end_cl = col_letter(source_layout['arr_end'])

    # --- Row 1: Fiscal month end + column totals ---
    ws.cell(row=1, column=1, value="Fiscal Month End:")
    ws.cell(row=1, column=2, value="=Control!C7")

    for section_key in ['arr', 'churn', 'downsell', 'upsell', 'new_biz']:
        start = layout[f'{section_key}_start']
        end = layout[f'{section_key}_end'] if section_key != 'arr' else layout['arr_end']
        for c in range(start, end + 1):
            cl = col_letter(c)
            ws.cell(row=1, column=c,
                    value=f"=SUM({cl}{first_data_row}:{cl}{last_data_row})")

    # --- Row 2: Quarter numbers ---
    ws.cell(row=2, column=layout['label'], value="Quarter")

    if target_granularity == 'quarterly':
        # Find first quarter from source monthly data
        ws.cell(row=2, column=layout['arr_start'],
                value=f"=INDEX('{source_sheet}'!${src_arr_start_cl}2:${src_arr_end_cl}2,"
                      f"MATCH(TRUE,INDEX('{source_sheet}'!${src_arr_start_cl}2:${src_arr_end_cl}2<>0,),0))")
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=2, column=layout['arr_start'] + i,
                    value=f"=IF({prev_cl}2=4,1,{prev_cl}2+1)")
    elif target_granularity == 'annual':
        # Annual: always Q4
        ws.cell(row=2, column=layout['arr_start'], value=4)
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=2, column=layout['arr_start'] + i,
                    value=f"={prev_cl}2")

    _copy_helper_rows_to_derived(ws, layout, 2)

    # --- Row 3: Year numbers ---
    ws.cell(row=3, column=layout['label'], value="Year")

    if target_granularity == 'quarterly':
        ws.cell(row=3, column=layout['arr_start'],
                value=f"=INDEX('{source_sheet}'!${src_arr_start_cl}3:${src_arr_end_cl}3,"
                      f"MATCH(TRUE,INDEX('{source_sheet}'!${src_arr_start_cl}3:${src_arr_end_cl}3<>0,),0))")
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            prev_q_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=3, column=layout['arr_start'] + i,
                    value=f"=IF({prev_q_cl}2=4,{prev_cl}3+1,{prev_cl}3)")
    elif target_granularity == 'annual':
        ws.cell(row=3, column=layout['arr_start'],
                value=f"=INDEX('{source_sheet}'!${src_arr_start_cl}3:${src_arr_end_cl}3,"
                      f"MATCH(TRUE,INDEX('{source_sheet}'!${src_arr_start_cl}3:${src_arr_end_cl}3<>0,),0))")
        for i in range(1, num_dates):
            prev_cl = col_letter(layout['arr_start'] + i - 1)
            ws.cell(row=3, column=layout['arr_start'] + i,
                    value=f"=IF({prev_cl}2=4,{prev_cl}3+1,{prev_cl}3)")

    _copy_helper_rows_to_derived(ws, layout, 3)

    # --- Row 5: Section headers ---
    ws.cell(row=5, column=layout['attr_start'], value="Customer Identifying Information")
    ws.cell(row=5, column=layout['arr_start'],
            value=f"{target_granularity.capitalize()} ARR by Date")
    ws.cell(row=5, column=layout['churn_start'], value="Churn?")
    ws.cell(row=5, column=layout['downsell_start'], value="Downsell?")
    ws.cell(row=5, column=layout['upsell_start'], value="Upsell?")
    ws.cell(row=5, column=layout['new_biz_start'], value="New Business Dollars?")

    # --- Row 6: Column headers ---
    ws.cell(row=6, column=layout['cust_id'], value="Customer ID")
    attr_names = list(config['attributes'].keys())
    for i, name in enumerate(attr_names):
        ws.cell(row=6, column=layout['attr_start'] + i, value=name)

    cohort_name = f"{target_granularity.capitalize()} Cohort"
    ws.cell(row=6, column=layout['cohort'], value=cohort_name)
    ws.cell(row=6, column=layout['rank'], value="Customer Rank #")

    # Date headers: formula-based labels
    for i in range(num_dates):
        c = layout['arr_start'] + i
        cl = col_letter(c)
        if target_granularity == 'quarterly':
            ws.cell(row=6, column=c, value=f'="Q"&{cl}2&"\'"&RIGHT({cl}3,2)')
        elif target_granularity == 'annual':
            ws.cell(row=6, column=c, value=f'="FY"&"\'"&RIGHT({cl}3,2)')

    # Derived section date headers
    for i in range(layout['num_derived']):
        arr_ref_col = layout['arr_start'] + yoy_offset + i
        arr_ref_cl = col_letter(arr_ref_col)

        churn_col = layout['churn_start'] + i
        ws.cell(row=6, column=churn_col, value=f"={arr_ref_cl}6")

        downsell_col = layout['downsell_start'] + i
        ws.cell(row=6, column=downsell_col, value=f"={col_letter(churn_col)}6")

        upsell_col = layout['upsell_start'] + i
        ws.cell(row=6, column=upsell_col, value=f"={col_letter(downsell_col)}6")

        new_biz_col = layout['new_biz_start'] + i
        ws.cell(row=6, column=new_biz_col, value=f"={col_letter(upsell_col)}6")

    # --- Customer data rows ---
    arr_start_cl = col_letter(layout['arr_start'])
    arr_end_cl = col_letter(layout['arr_end'])

    for idx, cust_id in enumerate(unique_customers):
        row = first_data_row + idx
        cust_id_cl = col_letter(layout['cust_id'])

        # Customer ID: reference source sheet
        ws.cell(row=row, column=layout['cust_id'],
                value=f"='{source_sheet}'!{col_letter(source_layout['cust_id'])}{row}")

        # Attributes: reference source sheet
        for attr_idx in range(num_attrs):
            src_col = source_layout['attr_start'] + attr_idx
            ws.cell(row=row, column=layout['attr_start'] + attr_idx,
                    value=f"='{source_sheet}'!{col_letter(src_col)}{row}")

        # Cohort (first non-zero period)
        ws.cell(row=row, column=layout['cohort'],
                value=f'=IFERROR(INDEX(${arr_start_cl}$6:${arr_end_cl}$6,'
                      f'MATCH(TRUE,INDEX({arr_start_cl}{row}:{arr_end_cl}{row}<>0,0),0)),"n.a.")')

        # Rank (by last-period ARR)
        ws.cell(row=row, column=layout['rank'],
                value=f"=RANK({arr_end_cl}{row},"
                      f"${arr_end_cl}${first_data_row}:${arr_end_cl}${last_data_row})")

        # ARR: SUMIFS across source row matching Quarter & Year
        for i in range(num_dates):
            c = layout['arr_start'] + i
            cl = col_letter(c)
            ws.cell(row=row, column=c,
                    value=f"=SUMIFS('{source_sheet}'!${src_arr_start_cl}{row}:${src_arr_end_cl}{row},"
                          f"'{source_sheet}'!${src_arr_start_cl}$3:${src_arr_end_cl}$3,{cl}$3,"
                          f"'{source_sheet}'!${src_arr_start_cl}$2:${src_arr_end_cl}$2,{cl}$2)")

        # Derived sections (same logic as base tab)
        for i in range(layout['num_derived']):
            prior_cl = col_letter(layout['arr_start'] + i)
            curr_cl = col_letter(layout['arr_start'] + yoy_offset + i)
            prior_ref = f"{prior_cl}{row}"
            curr_ref = f"{curr_cl}{row}"

            ws.cell(row=row, column=layout['churn_start'] + i,
                    value=f"=IF(AND({curr_ref}=0,{prior_ref}>0),-{prior_ref},0)")
            ws.cell(row=row, column=layout['downsell_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}>0,{curr_ref}<{prior_ref}),"
                          f"{curr_ref}-{prior_ref},0)")
            ws.cell(row=row, column=layout['upsell_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}>0,{curr_ref}>{prior_ref}),"
                          f"{curr_ref}-{prior_ref},0)")
            ws.cell(row=row, column=layout['new_biz_start'] + i,
                    value=f"=IF(AND({curr_ref}>0,{prior_ref}=0),{curr_ref},0)")

    return sheet_name, layout, first_data_row, last_data_row


# ---------------------------------------------------------------------------
# Helper: Copy quarter/year rows to derived sections
# ---------------------------------------------------------------------------

def _copy_helper_rows_to_derived(ws, layout, helper_row):
    """
    Copy the Quarter (row 2) or Year (row 3) helper row from the ARR section
    to the Churn/Downsell/Upsell/New Business sections.

    Each derived section references the previous section's values.
    """
    yoy = layout['yoy_offset']
    num_derived = layout['num_derived']

    for i in range(num_derived):
        # Churn references ARR (offset by yoy_offset)
        arr_src_col = layout['arr_start'] + yoy + i
        churn_col = layout['churn_start'] + i
        ws.cell(row=helper_row, column=churn_col,
                value=f"={col_letter(arr_src_col)}{helper_row}")

        # Downsell references Churn
        downsell_col = layout['downsell_start'] + i
        ws.cell(row=helper_row, column=downsell_col,
                value=f"={col_letter(churn_col)}{helper_row}")

        # Upsell references Downsell
        upsell_col = layout['upsell_start'] + i
        ws.cell(row=helper_row, column=upsell_col,
                value=f"={col_letter(downsell_col)}{helper_row}")

        # New Biz references Upsell
        new_biz_col = layout['new_biz_start'] + i
        ws.cell(row=helper_row, column=new_biz_col,
                value=f"={col_letter(upsell_col)}{helper_row}")
