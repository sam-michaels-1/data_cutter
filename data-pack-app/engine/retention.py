"""
Retention tab generators.

Reference structure (Annual Retention from built-out file):
  Block 1 (rows 5-22): "Total Business"
  Block 2 (rows 24-41): "Software & Services"

  Within each block (18 data rows + 1 gap = 19):
    +0: Title row
    +1: Section sub-headers (Net Retention, Customer Retention, ARR/Customer)
    +2: Column headers (filter labels + dates)
    +3: BoP ARR / BoP Customers / BoP ARR/Cust
    +4: (-) Churn / (-) Churned Customers / (-) Churned ARR/Cust
    +5: (-) Downsell / Retained Customers / (+/-) Upsell/Cross-sell ARR/Cust
    +6: (+) Upsell / (+) New Logo / Retained ARR/Cust
    +7: Retained ARR / EoP Customers / (+) New Logo ARR/Cust
    +8: (+) New Logo / % Growth / EoP ARR/Cust
    +9: EoP ARR / (blank) / % Growth ARR/Cust.
    +10: % Growth / (blank) / (blank)
    +11: Check / Check / (blank)
    +12: (blank)
    +13: % Lost-Only Retention / Logo Retention / New Logo vs Churn
    +14: % Punitive Retention / New Logo % of BoP / New Logo vs Retained
    +15: % Net Retention (bold)
    +16: % New Logo % of BoP
    +17: % New Logo Growth
    +18: (gap row before next block)
"""
from openpyxl.utils import get_column_letter
from .utils import col_letter, get_yoy_offset


BLOCK_HEIGHT = 19  # 18 data rows + 1 gap row between blocks


def generate_retention_tab(wb, config, clean_sheet_name, clean_layout,
                           first_data_row, last_data_row, granularity,
                           filter_blocks):
    """Generate a retention analysis tab."""
    sheet_name = f"{granularity.capitalize()} Retention"
    ws = wb.create_sheet(sheet_name)

    num_derived = clean_layout['num_derived']
    num_attrs = clean_layout['num_attrs']
    attr_names = list(config['attributes'].keys())
    yoy_offset = clean_layout['yoy_offset']

    # Column layout:
    # B-E: filter criteria (attrs + cohort)
    # G: Section 1 label, H..H+n-1: Section 1 data
    # gap col
    # M_label: Section 2 label, N..N+n-1: Section 2 data
    # gap col
    # S_label: Section 3 label, T..T+n-1: Section 3 data
    filter_start_col = 2  # B
    cohort_fc = filter_start_col + num_attrs  # after attrs

    s1_label = cohort_fc + 2  # G (skip one gap col)
    s1_start = s1_label + 1   # H
    s1_end = s1_start + num_derived - 1

    s2_label = s1_end + 2     # M (skip one gap col)
    s2_start = s2_label + 1   # N
    s2_end = s2_start + num_derived - 1

    s3_label = s2_end + 2     # S (skip one gap col)
    s3_start = s3_label + 1   # T
    s3_end = s3_start + num_derived - 1

    # --- Row 1: Check summary ---
    ws.cell(row=1, column=1, value="Check Summary")
    check_refs = []
    for block_idx in range(len(filter_blocks)):
        block_start = 5 + block_idx * BLOCK_HEIGHT
        check_row = block_start + 11  # Check row
        check_refs.append(f"{col_letter(s1_start)}{check_row}:{col_letter(s1_end)}{check_row}")
        check_refs.append(f"{col_letter(s2_start)}{check_row}:{col_letter(s2_end)}{check_row}")
    ws.cell(row=1, column=2, value=f"=SUM({','.join(check_refs)})")

    # --- Row 3: Units ---
    ws.cell(row=3, column=1, value="Units")
    ws.cell(row=3, column=filter_start_col, value="=Control!$C$4")
    units_cell = f"${col_letter(filter_start_col)}$3"

    # --- Generate each block ---
    for block_idx, block in enumerate(filter_blocks):
        block_start = 5 + block_idx * BLOCK_HEIGHT
        _write_retention_block(
            ws, block_start, block, config,
            clean_sheet_name, clean_layout,
            first_data_row, last_data_row,
            s1_label, s1_start, s1_end,
            s2_label, s2_start, s2_end,
            s3_label, s3_start, s3_end,
            filter_start_col, cohort_fc,
            num_derived, num_attrs, attr_names,
            units_cell, yoy_offset, granularity
        )

    return sheet_name


def _write_retention_block(ws, start, block, config,
                           clean_sheet, clean_layout,
                           cdr_first, cdr_last,
                           s1_label, s1_start, s1_end,
                           s2_label, s2_start, s2_end,
                           s3_label, s3_start, s3_end,
                           filter_start, cohort_fc,
                           num_derived, num_attrs, attr_names,
                           units_cell, yoy_offset, granularity):
    """Write one retention block (18 rows + 1 gap)."""
    title = block['title']
    filters = block['filters']

    # Row positions within block
    r_title     = start + 0
    r_sections  = start + 1
    r_header    = start + 2
    r_bop       = start + 3
    r_churn     = start + 4
    r_downsell  = start + 5
    r_upsell    = start + 6
    r_retained  = start + 7
    r_new_logo  = start + 8
    r_eop       = start + 9
    r_growth    = start + 10
    r_check     = start + 11
    # blank       = start + 12
    r_lost_ret  = start + 13
    r_punit_ret = start + 14
    r_net_ret   = start + 15
    r_nl_pct    = start + 16
    r_nl_growth = start + 17

    # --- Title row ---
    ws.cell(row=r_title, column=s1_label,
            value=f"{title} {granularity.capitalize()} Retention Analysis")

    # --- Section sub-headers ---
    ws.cell(row=r_sections, column=s1_label, value="Net Retention Analysis")
    ws.cell(row=r_sections, column=s2_label, value="Customer Retention Analysis")
    metric_label = "ARR" if config.get("data_type", "arr") == "arr" else "Revenue"
    ws.cell(row=r_sections, column=s3_label, value=f"{metric_label} / Customer")

    # --- Column header row ---
    # Filter labels
    for attr_idx, attr_name in enumerate(attr_names):
        ws.cell(row=r_header, column=filter_start + attr_idx, value=attr_name)
    ws.cell(row=r_header, column=cohort_fc, value="Cohort")

    # Date headers from clean data's churn section
    for i in range(num_derived):
        churn_col = clean_layout['churn_start'] + i
        churn_cl = col_letter(churn_col)
        # Section 1 dates
        ws.cell(row=r_header, column=s1_start + i,
                value=f"='{clean_sheet}'!{churn_cl}$6")
        # Section 2 dates reference section 1
        ws.cell(row=r_header, column=s2_start + i,
                value=f"={col_letter(s1_start + i)}{r_header}")
        # Section 3 dates reference section 2
        ws.cell(row=r_header, column=s3_start + i,
                value=f"={col_letter(s2_start + i)}{r_header}")

    # --- Filter values (on BoP row, cascaded down) ---
    all_data_rows = [r_bop, r_churn, r_downsell, r_upsell, r_retained,
                     r_new_logo, r_eop, r_growth, r_check,
                     r_lost_ret, r_punit_ret, r_net_ret, r_nl_pct, r_nl_growth]
    for attr_idx, attr_name in enumerate(attr_names):
        fc = filter_start + attr_idx
        filter_val = filters.get(attr_name, "<>")
        ws.cell(row=r_bop, column=fc, value=filter_val)
        for r in all_data_rows[1:]:  # skip r_bop itself
            ws.cell(row=r, column=fc, value=f"={col_letter(fc)}{r_bop}")

    # Cohort filter
    ws.cell(row=r_bop, column=cohort_fc, value=filters.get("Cohort", "<>"))
    for r in all_data_rows[1:]:
        ws.cell(row=r, column=cohort_fc, value=f"={col_letter(cohort_fc)}{r_bop}")

    # --- Helper: build SUMIFS criteria ---
    def _criteria(row):
        """Build criteria pairs string for SUMIFS/COUNTIFS."""
        parts = []
        for attr_idx in range(num_attrs):
            cc = clean_layout['attr_start'] + attr_idx
            parts.append(
                f"'{clean_sheet}'!${col_letter(cc)}${cdr_first}"
                f":${col_letter(cc)}${cdr_last},"
                f"${col_letter(filter_start + attr_idx)}{row}")
        # Cohort
        cc = clean_layout['cohort']
        parts.append(
            f"'{clean_sheet}'!${col_letter(cc)}${cdr_first}"
            f":${col_letter(cc)}${cdr_last},"
            f"${col_letter(cohort_fc)}{row}")
        return ','.join(parts)

    def _sumifs(clean_col_num, row):
        """Build a full SUMIFS formula divided by units."""
        cl = col_letter(clean_col_num)
        rng = f"'{clean_sheet}'!{cl}${cdr_first}:{cl}${cdr_last}"
        return f"=SUMIFS({rng},{_criteria(row)})/{units_cell}"

    def _countifs_nonzero(clean_col_num, row):
        """Build a COUNTIFS formula counting non-zero values."""
        cl = col_letter(clean_col_num)
        rng = f"'{clean_sheet}'!{cl}${cdr_first}:{cl}${cdr_last}"
        return f'=COUNTIFS({rng},"<>"&0,{_criteria(row)})'

    # ===== SECTION 1: Net Retention Analysis (ARR) =====
    # Row labels
    ws.cell(row=r_bop,       column=s1_label, value=f"BoP {metric_label}")
    ws.cell(row=r_churn,     column=s1_label, value="(-) Churn")
    ws.cell(row=r_downsell,  column=s1_label, value="(-) Downsell")
    ws.cell(row=r_upsell,    column=s1_label, value="(+) Upsell / Cross-sell")
    ws.cell(row=r_retained,  column=s1_label, value=f"Retained {metric_label}")
    ws.cell(row=r_new_logo,  column=s1_label, value="(+) New Logo")
    ws.cell(row=r_eop,       column=s1_label, value=f"EoP {metric_label}")
    ws.cell(row=r_growth,    column=s1_label, value="% Growth")
    ws.cell(row=r_check,     column=s1_label, value="Check")
    ws.cell(row=r_lost_ret,  column=s1_label, value="% Lost-Only Retention")
    ws.cell(row=r_punit_ret, column=s1_label, value="% Punitive Retention")
    ws.cell(row=r_net_ret,   column=s1_label, value="% Net Retention")
    ws.cell(row=r_nl_pct,    column=s1_label, value="% New Logo % of BoP")
    ws.cell(row=r_nl_growth, column=s1_label, value="% New Logo Growth")

    for i in range(num_derived):
        dc = s1_start + i
        dcl = col_letter(dc)

        # BoP ARR
        bop_col = clean_layout['arr_start'] + i
        ws.cell(row=r_bop, column=dc, value=_sumifs(bop_col, r_bop))

        # Churn
        churn_col = clean_layout['churn_start'] + i
        ws.cell(row=r_churn, column=dc, value=_sumifs(churn_col, r_churn))

        # Downsell
        ds_col = clean_layout['downsell_start'] + i
        ws.cell(row=r_downsell, column=dc, value=_sumifs(ds_col, r_downsell))

        # Upsell
        up_col = clean_layout['upsell_start'] + i
        ws.cell(row=r_upsell, column=dc, value=_sumifs(up_col, r_upsell))

        # Retained = SUM(BoP:Upsell)
        ws.cell(row=r_retained, column=dc,
                value=f"=SUM({dcl}{r_bop}:{dcl}{r_upsell})")

        # New Logo
        nb_col = clean_layout['new_biz_start'] + i
        ws.cell(row=r_new_logo, column=dc, value=_sumifs(nb_col, r_new_logo))

        # EoP = Retained + New Logo
        ws.cell(row=r_eop, column=dc,
                value=f"=SUM({dcl}{r_retained}:{dcl}{r_new_logo})")

        # % Growth
        ws.cell(row=r_growth, column=dc,
                value=f"={dcl}{r_eop}/{dcl}{r_bop}-1")

        # Check: EoP*units - SUMIFS(EoP ARR col from clean data)
        eop_col = clean_layout['arr_start'] + yoy_offset + i
        eop_cl = col_letter(eop_col)
        eop_rng = f"'{clean_sheet}'!{eop_cl}${cdr_first}:{eop_cl}${cdr_last}"
        ws.cell(row=r_check, column=dc,
                value=f"={dcl}{r_eop}*{units_cell}-SUMIFS({eop_rng},{_criteria(r_check)})")

        # % Lost-Only Retention = (BoP + Churn) / BoP
        ws.cell(row=r_lost_ret, column=dc,
                value=f"=SUM({dcl}{r_bop}:{dcl}{r_churn})/{dcl}{r_bop}")

        # % Punitive Retention = (BoP + Churn + Downsell) / BoP
        ws.cell(row=r_punit_ret, column=dc,
                value=f"=SUM({dcl}{r_bop}:{dcl}{r_downsell})/{dcl}{r_bop}")

        # % Net Retention = Retained / BoP = (BoP + Churn + Downsell + Upsell) / BoP
        ws.cell(row=r_net_ret, column=dc,
                value=f"=SUM({dcl}{r_bop}:{dcl}{r_upsell})/{dcl}{r_bop}")

        # % New Logo % of BoP
        ws.cell(row=r_nl_pct, column=dc,
                value=f"={dcl}{r_new_logo}/{dcl}{r_bop}")

        # % New Logo Growth (YoY: compare to same period one year ago)
        if i >= yoy_offset:
            prior_cl = col_letter(s1_start + i - yoy_offset)
            ws.cell(row=r_nl_growth, column=dc,
                    value=f"={dcl}{r_new_logo}/{prior_cl}{r_new_logo}-1")

    # ===== SECTION 2: Customer Retention Analysis =====
    # Row labels for Section 2
    ws.cell(row=r_bop,       column=s2_label, value="BoP Customers")
    ws.cell(row=r_churn,     column=s2_label, value="(-) Churned Customers")
    ws.cell(row=r_downsell,  column=s2_label, value="Retained Customers")
    ws.cell(row=r_upsell,    column=s2_label, value="(+) New Logo")
    ws.cell(row=r_retained,  column=s2_label, value="EoP Customers")
    ws.cell(row=r_new_logo,  column=s2_label, value="% Growth")
    ws.cell(row=r_check,     column=s2_label, value="Check")
    ws.cell(row=r_lost_ret,  column=s2_label, value="Logo Retention")
    ws.cell(row=r_punit_ret, column=s2_label, value="New Logo % of BoP")

    for i in range(num_derived):
        dc = s2_start + i
        dcl = col_letter(dc)

        # BoP Customers (count non-zero ARR in BoP period)
        bop_col = clean_layout['arr_start'] + i
        ws.cell(row=r_bop, column=dc, value=_countifs_nonzero(bop_col, r_bop))

        # Churned Customers (negative count)
        churn_col = clean_layout['churn_start'] + i
        cl = col_letter(churn_col)
        rng = f"'{clean_sheet}'!{cl}${cdr_first}:{cl}${cdr_last}"
        ws.cell(row=r_churn, column=dc,
                value=f'=-(COUNTIFS({rng},"<>"&0,{_criteria(r_churn)}))')

        # Retained Customers = BoP + Churned
        ws.cell(row=r_downsell, column=dc,
                value=f"=SUM({dcl}{r_bop}:{dcl}{r_churn})")

        # New Logo customers (count non-zero new business)
        nb_col = clean_layout['new_biz_start'] + i
        ws.cell(row=r_upsell, column=dc, value=_countifs_nonzero(nb_col, r_upsell))

        # EoP Customers = Retained + New Logo
        ws.cell(row=r_retained, column=dc,
                value=f"=SUM({dcl}{r_downsell}:{dcl}{r_upsell})")

        # % Growth = EoP/BoP - 1
        ws.cell(row=r_new_logo, column=dc,
                value=f"={dcl}{r_retained}/{dcl}{r_bop}-1")

        # Check: EoP - COUNTIFS(EoP ARR col from clean data)
        eop_col = clean_layout['arr_start'] + yoy_offset + i
        eop_cl = col_letter(eop_col)
        eop_rng = f"'{clean_sheet}'!{eop_cl}${cdr_first}:{eop_cl}${cdr_last}"
        ws.cell(row=r_check, column=dc,
                value=f'={dcl}{r_retained}-COUNTIFS({eop_rng},"<>"&0,{_criteria(r_check)})')

        # Logo Retention = Retained / BoP
        ws.cell(row=r_lost_ret, column=dc,
                value=f"={dcl}{r_downsell}/{dcl}{r_bop}")

        # New Logo % of BoP
        ws.cell(row=r_punit_ret, column=dc,
                value=f"={dcl}{r_upsell}/{dcl}{r_bop}")

    # ===== SECTION 3: ARR / Customer =====
    # Row labels
    ws.cell(row=r_bop,       column=s3_label, value="BoP Customers")
    ws.cell(row=r_churn,     column=s3_label, value="(-) Churned Customers")
    ws.cell(row=r_downsell,  column=s3_label, value="(+/-) Upsell / Cross-sell")
    ws.cell(row=r_upsell,    column=s3_label, value="Retained Customers")
    ws.cell(row=r_retained,  column=s3_label, value="(+) New Logo")
    ws.cell(row=r_new_logo,  column=s3_label, value="EoP Customers")
    ws.cell(row=r_eop,       column=s3_label, value=f"% Growth {metric_label}/Cust.")
    ws.cell(row=r_lost_ret,  column=s3_label, value="New Logo vs Churn")
    ws.cell(row=r_punit_ret, column=s3_label, value="New Logo vs Retained")

    for i in range(num_derived):
        dc = s3_start + i
        dcl = col_letter(dc)
        s1_dcl = col_letter(s1_start + i)   # ARR column
        s2_dcl = col_letter(s2_start + i)   # Customer column

        # BoP ARR / Customer
        ws.cell(row=r_bop, column=dc,
                value=f'=IFERROR({s1_dcl}{r_bop}/{s2_dcl}{r_bop},"n.a.")')

        # (-) Churned Customers ARR/Cust = -Churn_ARR / Churned_Customers
        ws.cell(row=r_churn, column=dc,
                value=f'=IFERROR(-{s1_dcl}{r_churn}/{s2_dcl}{r_churn},"n.a.")')

        # (+/-) Upsell/Cross-sell per customer = (EoP ARR/Retained Cust) - SUM(BoP:Churn ARR/Cust)
        ws.cell(row=r_downsell, column=dc,
                value=f'=IFERROR({dcl}{r_upsell}-SUM({dcl}{r_bop}:{dcl}{r_churn}),"n.a.")')

        # Retained Customers ARR/Cust = Retained ARR / Retained Customers
        ws.cell(row=r_upsell, column=dc,
                value=f'=IFERROR({s1_dcl}{r_retained}/{s2_dcl}{r_downsell},"n.a.")')

        # (+) New Logo ARR/Cust = New Logo ARR / New Logo Customers
        ws.cell(row=r_retained, column=dc,
                value=f'=IFERROR({s1_dcl}{r_new_logo}/{s2_dcl}{r_upsell},"n.a.")')

        # EoP ARR / Customer = EoP ARR / EoP Customers
        ws.cell(row=r_new_logo, column=dc,
                value=f'=IFERROR({s1_dcl}{r_eop}/{s2_dcl}{r_retained},"n.a.")')

        # % Growth ARR/Cust = EoP ARR/Cust / BoP ARR/Cust - 1
        ws.cell(row=r_eop, column=dc,
                value=f'=IFERROR({dcl}{r_new_logo}/{dcl}{r_bop}-1,"n.a.")')

        # New Logo vs Churn = -(New Logo ARR/Cust / Churned ARR/Cust)
        ws.cell(row=r_lost_ret, column=dc,
                value=f'=IFERROR(-{dcl}{r_retained}/{dcl}{r_churn},"n.a.")')

        # New Logo vs Retained = New Logo ARR/Cust / Retained ARR/Cust
        ws.cell(row=r_punit_ret, column=dc,
                value=f'=IFERROR({dcl}{r_retained}/{dcl}{r_upsell},"n.a.")')

    return
