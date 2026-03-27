"""
Inspect the 'Data Pack AI Clean Version-Built Out.xlsx' file in detail.
Prints sheet structure, formatting, merged cells, frozen panes, column widths,
and cell-level details (values, formulas, fonts, fills, alignment, number formats, borders).
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys

FILE_PATH = "/Users/sammichaels/Desktop/Coding/Data Cutter/Data Pack AI Clean Version-Built Out.xlsx"
OUTPUT_PATH = "/Users/sammichaels/Desktop/Coding/Data Cutter/inspection_output.txt"

# Tabs to skip detailed cell inspection (too large)
SKIP_DETAIL_TABS = {"Raw Data", "Clean Data", "raw data", "clean data"}

# Tabs needing extended row inspection
EXTENDED_TABS = {
    "Annual Cohort": 80,
    "Annual Retention": 80,
    "Monthly Retention": 80,
    "Annual Top Customer Analysis": 80,
}

def color_to_str(color):
    """Convert an openpyxl Color object to a readable string."""
    if color is None:
        return "None"
    if color.type == "rgb" and color.rgb:
        return f"RGB({color.rgb})"
    if color.type == "theme":
        return f"Theme({color.theme}, tint={color.tint})"
    if color.type == "indexed":
        return f"Indexed({color.indexed})"
    return str(color)

def border_side_str(side):
    if side is None or side.style is None:
        return "none"
    return f"{side.style}({color_to_str(side.color)})"

def fmt_cell(cell):
    """Return a dict of formatting details for a cell."""
    info = {}
    info["value"] = cell.value
    info["data_type"] = cell.data_type  # 's'=string, 'n'=number, 'f'=formula, etc.

    # Font
    f = cell.font
    if f:
        info["font"] = {
            "name": f.name,
            "size": f.size,
            "bold": f.bold,
            "italic": f.italic,
            "underline": f.underline,
            "color": color_to_str(f.color),
        }

    # Fill
    fl = cell.fill
    if fl:
        info["fill"] = {
            "type": fl.fill_type,
            "fgColor": color_to_str(fl.fgColor) if fl.fgColor else "None",
            "bgColor": color_to_str(fl.bgColor) if fl.bgColor else "None",
        }

    # Alignment
    a = cell.alignment
    if a:
        info["alignment"] = {
            "horizontal": a.horizontal,
            "vertical": a.vertical,
            "wrap_text": a.wrap_text,
            "indent": a.indent,
        }

    # Number format
    info["number_format"] = cell.number_format

    # Border
    b = cell.border
    if b:
        info["border"] = {
            "left": border_side_str(b.left),
            "right": border_side_str(b.right),
            "top": border_side_str(b.top),
            "bottom": border_side_str(b.bottom),
        }

    return info


def inspect_workbook(out):
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False)

    out.write(f"=== WORKBOOK: {FILE_PATH} ===\n")
    out.write(f"Sheet names: {wb.sheetnames}\n")
    out.write(f"Number of sheets: {len(wb.sheetnames)}\n\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n{'='*100}\n")
        out.write(f"SHEET: '{sheet_name}'\n")
        out.write(f"{'='*100}\n")

        # Dimensions
        out.write(f"  Dimensions: {ws.dimensions}\n")
        out.write(f"  Min row: {ws.min_row}, Max row: {ws.max_row}\n")
        out.write(f"  Min col: {ws.min_column}, Max col: {ws.max_column}\n")

        # Merged cells
        merged = list(ws.merged_cells.ranges)
        out.write(f"  Merged cells ({len(merged)}): {merged[:50]}\n")
        if len(merged) > 50:
            out.write(f"    ... and {len(merged)-50} more\n")

        # Frozen panes
        out.write(f"  Freeze panes: {ws.freeze_panes}\n")

        # Sheet properties
        out.write(f"  Sheet state: {ws.sheet_state}\n")
        if hasattr(ws, 'sheet_properties'):
            sp = ws.sheet_properties
            out.write(f"  Tab color: {sp.tabColor}\n")

        # Column widths
        out.write(f"  Column widths/dimensions:\n")
        col_dims = dict(ws.column_dimensions)
        for col_letter in sorted(col_dims.keys(), key=lambda x: (len(x), x)):
            cd = col_dims[col_letter]
            out.write(f"    Col {col_letter}: width={cd.width}, hidden={cd.hidden}, "
                      f"bestFit={cd.bestFit}, auto_size={cd.auto_size}\n")

        # Row heights
        out.write(f"  Row heights (non-default):\n")
        row_dims = dict(ws.row_dimensions)
        for row_num in sorted(row_dims.keys()):
            rd = row_dims[row_num]
            if rd.height is not None:
                out.write(f"    Row {row_num}: height={rd.height}, hidden={rd.hidden}\n")

        # Skip detailed cell data for raw/clean data tabs
        if sheet_name in SKIP_DETAIL_TABS:
            out.write(f"\n  [SKIPPING detailed cell inspection for '{sheet_name}' - too large]\n")
            # Still show first 5 rows as a sample
            out.write(f"  Sample (first 5 rows):\n")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)), start=1):
                for cell in row:
                    if cell.value is not None:
                        out.write(f"    [{cell.coordinate}] = {repr(cell.value)}\n")
            continue

        # Determine how many rows to inspect
        max_rows = EXTENDED_TABS.get(sheet_name, 50)
        actual_max = min(max_rows, ws.max_row) if ws.max_row else 0

        out.write(f"\n  --- CELL DETAILS (rows 1-{actual_max}) ---\n")

        for row_idx in range(1, actual_max + 1):
            row_has_data = False
            row_output = []
            for col_idx in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None or (cell.fill and cell.fill.fill_type and cell.fill.fill_type != "none"):
                    row_has_data = True
                    ci = fmt_cell(cell)
                    col_letter = get_column_letter(col_idx)
                    row_output.append(f"    [{col_letter}{row_idx}] {ci}")

            if row_has_data:
                out.write(f"\n  Row {row_idx}:\n")
                for line in row_output:
                    out.write(line + "\n")

        # For specific tabs, also show extended regions
        if sheet_name == "Annual Cohort":
            out.write(f"\n  --- EXTENDED REGION: Rows 10-20, Cols A-AZ ---\n")
            for row_idx in range(10, min(21, (ws.max_row or 0) + 1)):
                for col_idx in range(1, min(53, (ws.max_column or 0) + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        ci = fmt_cell(cell)
                        col_letter = get_column_letter(col_idx)
                        out.write(f"    [{col_letter}{row_idx}] {ci}\n")

        if sheet_name in ("Annual Retention", "Monthly Retention"):
            out.write(f"\n  --- EXTENDED REGION: Rows 1-40, all columns ---\n")
            for row_idx in range(1, min(41, (ws.max_row or 0) + 1)):
                for col_idx in range(1, min(60, (ws.max_column or 0) + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        ci = fmt_cell(cell)
                        col_letter = get_column_letter(col_idx)
                        out.write(f"    [{col_letter}{row_idx}] {ci}\n")

    wb.close()
    out.write("\n\n=== INSPECTION COMPLETE ===\n")


if __name__ == "__main__":
    with open(OUTPUT_PATH, "w", encoding="utf-8") as out:
        inspect_workbook(out)
    print(f"Inspection output written to {OUTPUT_PATH}")
