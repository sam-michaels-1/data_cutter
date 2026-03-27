"""
Auto-detection service.

Reads the chosen sheet from an uploaded workbook and infers:
  - Which column is the Date column
  - Which column is the Customer ID column
  - Which column is the ARR / value column
  - Which columns are potential attribute/identifier columns
  - An appropriate scale factor (1, 1_000, or 1_000_000)
  - Total data row count
"""

import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Number of sample rows to examine for type inference
SAMPLE_ROWS = 50


def detect_columns(filepath: str, sheet_name: str) -> dict:
    """
    Read the given sheet and return a detection result dict.

    Returns
    -------
    {
        "columns": [{"letter", "header", "sample_values"}, ...],
        "detected_mapping": {"date_col", "customer_id_col", "arr_col", "attribute_cols"},
        "row_count": int,
        "auto_scale_factor": int,
    }
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # --- Read headers (row 1) and sample data rows ---
    rows_iter = ws.iter_rows(min_row=1, values_only=False)
    header_row = next(rows_iter)
    headers = []
    for cell in header_row:
        if cell.value is None:
            break
        headers.append({
            "col_num": cell.column,
            "letter": get_column_letter(cell.column),
            "header": str(cell.value).strip(),
        })

    num_cols = len(headers)
    if num_cols == 0:
        wb.close()
        raise ValueError("No headers found in row 1 of the selected sheet.")

    # Collect sample values per column (first SAMPLE_ROWS data rows)
    col_samples: dict[int, list] = {h["col_num"]: [] for h in headers}
    total_data_rows = 0

    for row in rows_iter:
        # Stop counting if the first cell is empty (end of data)
        if row[0].value is None:
            break
        total_data_rows += 1
        if total_data_rows <= SAMPLE_ROWS:
            for h in headers:
                idx = h["col_num"] - 1  # 0-based index in row tuple
                if idx < len(row):
                    col_samples[h["col_num"]].append(row[idx].value)

    # Continue counting remaining rows
    for row in rows_iter:
        if row[0].value is None:
            break
        total_data_rows += 1

    wb.close()

    # --- Classify each column ---
    date_col = None
    customer_id_col = None
    arr_col = None
    attribute_cols = []

    col_stats = {}
    for h in headers:
        cn = h["col_num"]
        vals = col_samples[cn]
        non_none = [v for v in vals if v is not None]
        if not non_none:
            col_stats[cn] = {"type": "empty"}
            continue

        # Check date-ness
        date_count = sum(1 for v in non_none if isinstance(v, (datetime.datetime, datetime.date)))
        date_ratio = date_count / len(non_none) if non_none else 0

        # Check numeric-ness (includes strings that look like numbers)
        def _is_numeric(v):
            if isinstance(v, bool):
                return False
            if isinstance(v, (int, float)):
                return True
            if isinstance(v, str):
                try:
                    float(v.replace(',', ''))
                    return True
                except (ValueError, AttributeError):
                    return False
            return False

        num_count = sum(1 for v in non_none if _is_numeric(v))
        num_ratio = num_count / len(non_none) if non_none else 0

        # Check string-ness
        str_count = sum(1 for v in non_none if isinstance(v, str))
        str_ratio = str_count / len(non_none) if non_none else 0

        # Unique count
        try:
            unique_count = len(set(str(v) for v in non_none))
        except Exception:
            unique_count = len(non_none)

        # Max abs value for numeric columns
        max_abs = 0
        if num_ratio > 0.5:
            nums = []
            for v in non_none:
                try:
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        nums.append(abs(v))
                    elif isinstance(v, str):
                        nums.append(abs(float(v.replace(',', ''))))
                except (ValueError, TypeError):
                    pass
            max_abs = max(nums) if nums else 0

        col_stats[cn] = {
            "date_ratio": date_ratio,
            "num_ratio": num_ratio,
            "str_ratio": str_ratio,
            "unique_count": unique_count,
            "total_sampled": len(non_none),
            "max_abs": max_abs,
        }

    # --- Detect date column: highest date_ratio > 0.8, leftmost wins ---
    for h in headers:
        cn = h["col_num"]
        s = col_stats.get(cn, {})
        if s.get("date_ratio", 0) > 0.8:
            date_col = h["letter"]
            break

    # --- Detect ARR column: numeric, highest max_abs ---
    best_arr_col = None
    best_max_abs = -1
    for h in headers:
        cn = h["col_num"]
        s = col_stats.get(cn, {})
        if h["letter"] == date_col:
            continue
        if s.get("num_ratio", 0) > 0.5 and s.get("max_abs", 0) > best_max_abs:
            best_max_abs = s["max_abs"]
            best_arr_col = h["letter"]
    arr_col = best_arr_col

    # --- Detect customer ID column: text, high cardinality ---
    best_cid_col = None
    best_cardinality = -1
    for h in headers:
        cn = h["col_num"]
        s = col_stats.get(cn, {})
        if h["letter"] in (date_col, arr_col):
            continue
        if s.get("str_ratio", 0) > 0.5:
            card = s.get("unique_count", 0)
            if card > best_cardinality:
                best_cardinality = card
                best_cid_col = h["letter"]
    customer_id_col = best_cid_col

    # --- Detect attribute columns: everything remaining with low cardinality ---
    for h in headers:
        cn = h["col_num"]
        s = col_stats.get(cn, {})
        if h["letter"] in (date_col, arr_col, customer_id_col):
            continue
        if s.get("type") == "empty":
            continue
        unique = s.get("unique_count", 0)
        # Low cardinality relative to sample size → likely an identifier/attribute
        if unique > 0 and unique < 100:
            attribute_cols.append({"header": h["header"], "letter": h["letter"]})

    # --- Auto scale factor ---
    # Estimate total portfolio value from sampled ARR column to pick a
    # sensible display scale.  We sum the sample, scale up to full row count.
    arr_total_estimate = 0
    if arr_col:
        arr_cn = next((h["col_num"] for h in headers if h["letter"] == arr_col), None)
        if arr_cn:
            sample_sum = 0
            for v in col_samples.get(arr_cn, []):
                try:
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        sample_sum += abs(v)
                    elif isinstance(v, str):
                        sample_sum += abs(float(v.replace(',', '')))
                except (ValueError, TypeError):
                    pass
            sampled_n = min(SAMPLE_ROWS, total_data_rows) or 1
            arr_total_estimate = sample_sum / sampled_n * total_data_rows

    auto_scale = 1
    if arr_total_estimate > 1_000_000_000:
        auto_scale = 1_000_000
    elif arr_total_estimate > 1_000_000:
        auto_scale = 1_000

    # --- Build column info for the frontend ---
    column_info = []
    for h in headers:
        cn = h["col_num"]
        samples = col_samples.get(cn, [])
        sample_strs = [str(v) for v in samples[:5] if v is not None]
        column_info.append({
            "letter": h["letter"],
            "header": h["header"],
            "sample_values": sample_strs,
        })

    return {
        "columns": column_info,
        "detected_mapping": {
            "date_col": date_col or headers[0]["letter"],
            "customer_id_col": customer_id_col or (headers[1]["letter"] if len(headers) > 1 else headers[0]["letter"]),
            "arr_col": arr_col or (headers[-1]["letter"] if headers else "A"),
            "attribute_cols": attribute_cols,
        },
        "row_count": total_data_rows,
        "auto_scale_factor": auto_scale,
    }
