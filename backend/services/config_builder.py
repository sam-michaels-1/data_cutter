"""
Translate wizard selections into the engine's config dictionary.

The engine (data-pack-app/engine/generator.py) expects a specific config dict
structure. This module builds that dict from the frontend's GenerateRequest.
"""

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def _detect_raw_data_frequency(filepath: str, sheet_name: str, date_col: str) -> str:
    """
    Auto-detect the raw data frequency by examining the date intervals.
    Returns 'monthly', 'quarterly', or 'annual'.
    """
    import datetime

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    col_idx = column_index_from_string(date_col)

    dates = []
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx,
                            values_only=True):
        val = row[0]
        if val is None:
            break
        if isinstance(val, (datetime.datetime, datetime.date)):
            dates.append(val if isinstance(val, datetime.date) else val.date())

    wb.close()

    if len(dates) < 2:
        return "annual"

    # Get sorted unique dates
    unique_dates = sorted(set(dates))
    if len(unique_dates) < 2:
        return "annual"

    # Average gap in days between consecutive unique dates
    gaps = [(unique_dates[i + 1] - unique_dates[i]).days
            for i in range(min(10, len(unique_dates) - 1))]
    avg_gap = sum(gaps) / len(gaps)

    if avg_gap < 45:
        return "monthly"
    elif avg_gap < 120:
        return "quarterly"
    else:
        return "annual"


def _get_unique_values(filepath: str, sheet_name: str, col_letter: str,
                       max_unique: int = 50) -> list[str]:
    """Read unique values from a column (for auto-generating filter breakouts)."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    col_idx = column_index_from_string(col_letter)

    values = set()
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx,
                            values_only=True):
        val = row[0]
        if val is None:
            break
        val_str = str(val).strip()
        if val_str:
            values.add(val_str)
        if len(values) >= max_unique:
            break

    wb.close()
    return sorted(values)


def build_engine_config(
    filepath: str,
    session_id: str,
    sheet_name: str,
    data_type: str,
    date_col: str,
    customer_id_col: str,
    arr_col: str,
    attributes: list[dict],          # [{"display_name": ..., "letter": ...}, ...]
    output_granularities: list[str],  # ["monthly", "quarterly", "annual"]
    fiscal_year_end_month: int,
    row_count: int,
    scale_factor: int,
) -> dict:
    """
    Build the config dict the engine expects.

    Parameters match the GenerateRequest schema fields.
    """
    # Build ordered attributes dict
    attrs = OrderedDict()
    for attr in attributes:
        attrs[attr["display_name"]] = attr["letter"]

    # Detect raw data frequency (this becomes time_granularity for the engine)
    raw_freq = _detect_raw_data_frequency(filepath, sheet_name, date_col)

    # Auto-generate filter breakouts from the first attribute
    filter_breakouts = []
    if attributes:
        first_attr = attributes[0]
        unique_vals = _get_unique_values(
            filepath, sheet_name, first_attr["letter"])
        for val in unique_vals:
            filter_breakouts.append({
                "title": val,
                "filters": {first_attr["display_name"]: val},
            })

    return {
        "raw_data_sheet": sheet_name,
        "raw_data_first_row": 2,
        "raw_data_last_row": row_count + 1,  # +1 because row 1 is the header
        "customer_id_col": customer_id_col,
        "date_col": date_col,
        "arr_col": arr_col,
        "attributes": attrs,
        "time_granularity": raw_freq,
        "output_granularities": output_granularities,
        "fiscal_year_end_month": fiscal_year_end_month,
        "scale_factor": scale_factor,
        "filter_breakouts": filter_breakouts,
        "data_type": data_type,
    }
