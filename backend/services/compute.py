"""
Dashboard computation engine.

Reads raw uploaded Excel data and computes all dashboard metrics using pandas,
mirroring the logic in the Excel formula engine (clean_data.py, retention.py,
cohort.py, top_customers.py) but producing actual computed values for JSON.
"""
from __future__ import annotations

import datetime
import json
import os
from typing import Optional, Tuple, Dict

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def load_config(session_dir: str) -> dict:
    """Load the saved engine config from the session directory."""
    config_path = os.path.join(session_dir, "config.json")
    with open(config_path, "r") as f:
        return json.load(f)


def _read_raw_data(filepath: str, config: dict) -> pd.DataFrame:
    """Read raw Excel data into a pandas DataFrame with proper column typing.
    Also reads attribute columns from config['attributes'] if present.
    """
    sheet_name = config["raw_data_sheet"]
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    date_idx = column_index_from_string(config["date_col"]) - 1
    cust_idx = column_index_from_string(config["customer_id_col"]) - 1
    arr_idx = column_index_from_string(config["arr_col"]) - 1

    # Build attribute column indices: {display_name: col_index}
    attr_cols = {}
    for name, letter in config.get("attributes", {}).items():
        attr_cols[name] = column_index_from_string(letter) - 1

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[date_idx]
        cust_val = row[cust_idx]
        arr_val = row[arr_idx]

        if date_val is None or cust_val is None:
            continue

        # Normalize date
        if isinstance(date_val, datetime.datetime):
            date_val = date_val.date()
        elif not isinstance(date_val, datetime.date):
            continue

        # Normalize ARR
        try:
            arr_val = float(arr_val) if arr_val is not None else 0.0
        except (ValueError, TypeError):
            arr_val = 0.0

        record = {
            "date": date_val,
            "customer_id": str(cust_val).strip(),
            "arr": arr_val,
        }

        # Read attribute values
        for name, col_idx in attr_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[name] = str(val).strip() if val is not None else ""

        rows.append(record)

    wb.close()
    return pd.DataFrame(rows)


def _assign_fiscal_period(date: datetime.date, fy_month: int) -> tuple[int, int]:
    """
    Assign fiscal quarter and year to a date.
    fy_month = fiscal year end month (e.g., 12 for Dec).
    """
    month = date.month
    if month > fy_month:
        fiscal_year = date.year + 1
    else:
        fiscal_year = date.year
    fiscal_quarter = int(((month - (fy_month + 1)) % 12) / 3) + 1
    return fiscal_year, fiscal_quarter


def _aggregate_to_granularity(df: pd.DataFrame, granularity: str,
                               fy_month: int) -> tuple[pd.DataFrame, dict]:
    """
    Aggregate raw data to the requested granularity.
    Returns:
        (agg_df, period_date_map) where agg_df has columns:
        customer_id, period_label, period_sort, arr
        and period_date_map maps period_label -> max date (ISO string).
    """
    # Add fiscal year/quarter
    fy_data = df["date"].apply(lambda d: _assign_fiscal_period(d, fy_month))
    df = df.copy()
    df["fy_year"] = fy_data.apply(lambda x: x[0])
    df["fy_quarter"] = fy_data.apply(lambda x: x[1])

    if granularity == "annual":
        # Only use records at the fiscal year-end month (ARR is a point-in-time
        # snapshot, not cumulative — summing interim months would inflate values).
        at_fy_end = df["date"].apply(lambda d: d.month == fy_month)
        valid_years = set(df[at_fy_end]["fy_year"].tolist())
        df_filtered = df[at_fy_end & df["fy_year"].isin(valid_years)]
        grouped = df_filtered.groupby(["customer_id", "fy_year"])["arr"].sum().reset_index()
        grouped["period_label"] = grouped["fy_year"].apply(lambda y: f"FY'{y % 100:02d}")
        grouped["period_sort"] = grouped["fy_year"]
        # Build period -> max date map
        date_map = (df_filtered.groupby("fy_year")["date"].max()
                    .reset_index()
                    .assign(period_label=lambda x: x["fy_year"].apply(lambda y: f"FY'{y % 100:02d}"))
                    .set_index("period_label")["date"]
                    .apply(lambda d: d.isoformat())
                    .to_dict())
        return grouped[["customer_id", "period_label", "period_sort", "arr"]], date_map

    elif granularity == "quarterly":
        # Only use records at the fiscal quarter-end month.
        df = df.copy()
        df["at_qend"] = df["date"].apply(lambda d: (fy_month - d.month) % 3 == 0)
        valid_qkeys = (df[df["at_qend"]][["fy_year", "fy_quarter"]]
                       .drop_duplicates())
        df_filtered = df[df["at_qend"]].merge(valid_qkeys, on=["fy_year", "fy_quarter"])
        grouped = df_filtered.groupby(["customer_id", "fy_year", "fy_quarter"])["arr"].sum().reset_index()
        grouped["period_label"] = grouped.apply(
            lambda r: f"Q{int(r['fy_quarter'])}'{int(r['fy_year']) % 100:02d}", axis=1)
        grouped["period_sort"] = grouped["fy_year"] * 10 + grouped["fy_quarter"]
        date_map = (df_filtered.assign(
            period_label=lambda x: x.apply(
                lambda r: f"Q{int(r['fy_quarter'])}'{int(r['fy_year']) % 100:02d}", axis=1))
            .groupby("period_label")["date"].max()
            .apply(lambda d: d.isoformat())
            .to_dict())
        return grouped[["customer_id", "period_label", "period_sort", "arr"]], date_map

    else:  # monthly
        grouped = df.groupby(["customer_id", "date"])["arr"].sum().reset_index()
        grouped["period_label"] = grouped["date"].apply(lambda d: d.strftime("%b '%y"))
        grouped["period_sort"] = grouped["date"].apply(lambda d: d.year * 100 + d.month)
        date_map = (grouped.groupby("period_label")["date"].max()
                    .apply(lambda d: d.isoformat())
                    .to_dict())
        return grouped[["customer_id", "period_label", "period_sort", "arr"]], date_map


def _build_arr_matrix(agg_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    Build customer x period ARR pivot table.
    Returns (pivot DataFrame, sorted period labels).
    """
    period_order = (agg_df[["period_label", "period_sort"]]
                    .drop_duplicates()
                    .sort_values("period_sort"))
    periods = period_order["period_label"].tolist()

    pivot = agg_df.pivot_table(
        index="customer_id",
        columns="period_label",
        values="arr",
        aggfunc="sum",
        fill_value=0.0,
    )
    pivot = pivot.reindex(columns=periods, fill_value=0.0)
    return pivot, periods


def _get_yoy_offset(granularity: str) -> int:
    return {"monthly": 12, "quarterly": 4, "annual": 1}[granularity]


def _compute_derived(pivot: pd.DataFrame, periods: list[str],
                     yoy_offset: int) -> dict:
    """
    Compute churn/downsell/upsell/new for each derived period pair.
    Returns dict with keys: churn, downsell, upsell, new_biz — each a DataFrame.
    """
    num_derived = len(periods) - yoy_offset
    if num_derived <= 0:
        return {"churn": pd.DataFrame(), "downsell": pd.DataFrame(),
                "upsell": pd.DataFrame(), "new_biz": pd.DataFrame()}

    derived_periods = periods[yoy_offset:]
    result = {k: pd.DataFrame(index=pivot.index, columns=derived_periods, dtype=float)
              for k in ["churn", "downsell", "upsell", "new_biz"]}

    for i in range(num_derived):
        prior_col = periods[i]
        curr_col = periods[yoy_offset + i]
        prior = pivot[prior_col]
        curr = pivot[curr_col]

        churn_mask = (curr == 0) & (prior > 0)
        result["churn"][curr_col] = np.where(churn_mask, -prior, 0.0)

        ds_mask = (curr > 0) & (prior > 0) & (curr < prior)
        result["downsell"][curr_col] = np.where(ds_mask, curr - prior, 0.0)

        up_mask = (curr > 0) & (prior > 0) & (curr > prior)
        result["upsell"][curr_col] = np.where(up_mask, curr - prior, 0.0)

        nb_mask = (curr > 0) & (prior == 0)
        result["new_biz"][curr_col] = np.where(nb_mask, curr, 0.0)

    return result


def _build_cohort_map(pivot: pd.DataFrame, periods: list[str]) -> dict:
    """Return {customer_id: first_non_zero_period_label}."""
    cohort_map = {}
    for cust in pivot.index:
        for p in periods:
            if pivot.loc[cust, p] > 0:
                cohort_map[cust] = p
                break
    return cohort_map


def _compute_attribute_options(raw_df: pd.DataFrame, attr_names: list[str]) -> list[dict]:
    """Return list of {name, values} with unique sorted values per attribute."""
    options = []
    for name in attr_names:
        if name in raw_df.columns:
            vals = sorted(
                v for v in raw_df[name].unique().tolist() if v and v != ""
            )
            options.append({"name": name, "values": vals})
    return options


def _build_attr_lookup(raw_df: pd.DataFrame, attr_names: list[str]) -> dict:
    """Return {customer_id: {attr_name: first_non_empty_value}}."""
    if not attr_names:
        return {}
    lookup = {}
    for cust_id, group in raw_df.groupby("customer_id"):
        attrs = {}
        for name in attr_names:
            if name in group.columns:
                non_empty = group[name][group[name] != ""]
                attrs[name] = non_empty.iloc[0] if len(non_empty) > 0 else ""
            else:
                attrs[name] = ""
        lookup[str(cust_id)] = attrs
    return lookup


def _compute_overview(pivot: pd.DataFrame, periods: list[str],
                      derived: dict, yoy_offset: int,
                      scale_factor: int, period_date_map: dict,
                      cohort_map: dict, attr_lookup: dict,
                      attr_names: list[str], top_n: int = 10) -> dict:
    """Compute overview dashboard data."""
    sf = scale_factor

    # ARR over time (total per period)
    arr_over_time = [float(pivot[p].sum() / sf) for p in periods]

    # ARR growth percentages — year-over-year (uses yoy_offset so quarterly/monthly
    # compare the same period in the prior year, not the prior period)
    arr_growth_pcts: list = []
    for i in range(len(arr_over_time)):
        if i < yoy_offset:
            arr_growth_pcts.append(None)
        else:
            prior = arr_over_time[i - yoy_offset]
            arr_growth_pcts.append(round(arr_over_time[i] / prior - 1, 4) if prior != 0 else None)

    # Latest period info
    latest_period_label = periods[-1] if periods else ""
    latest_period_date = period_date_map.get(latest_period_label, "")

    # Stats (latest derived period)
    derived_periods = periods[yoy_offset:]
    if not derived_periods:
        return {
            "periods": periods,
            "arr_over_time": [round(v, 2) for v in arr_over_time],
            "arr_growth_pcts": arr_growth_pcts,
            "waterfall": None,
            "stats": {
                "total_arr": arr_over_time[-1] if arr_over_time else 0,
                "customer_count": int((pivot[periods[-1]] > 0).sum()) if periods else 0,
                "net_retention_pct": None,
                "yoy_growth_pct": None,
                "lost_only_retention_pct": None,
                "punitive_retention_pct": None,
            },
            "top_customers": [],
            "latest_period_label": latest_period_label,
            "latest_period_date": latest_period_date,
        }

    latest = derived_periods[-1]
    prior_idx = periods.index(latest) - yoy_offset
    prior_period = periods[prior_idx]

    bop = float(pivot[prior_period].sum() / sf)
    churn_total = float(derived["churn"][latest].sum() / sf)
    downsell_total = float(derived["downsell"][latest].sum() / sf)
    upsell_total = float(derived["upsell"][latest].sum() / sf)
    new_logo_total = float(derived["new_biz"][latest].sum() / sf)
    retained = bop + churn_total + downsell_total + upsell_total
    eop = retained + new_logo_total

    total_arr = eop
    customer_count = int((pivot[periods[-1]] > 0).sum())

    net_retention_pct = retained / bop if bop != 0 else None
    lost_only_retention_pct = (bop + churn_total) / bop if bop != 0 else None
    punitive_retention_pct = (bop + churn_total + downsell_total) / bop if bop != 0 else None
    yoy_growth_pct = (eop / bop - 1) if bop != 0 else None

    waterfall = {
        "period_label": latest,
        "bop": round(bop, 2),
        "new_logo": round(new_logo_total, 2),
        "upsell": round(upsell_total, 2),
        "downsell": round(downsell_total, 2),
        "churn": round(churn_total, 2),
        "eop": round(eop, 2),
    }

    stats = {
        "total_arr": round(total_arr, 2),
        "customer_count": customer_count,
        "net_retention_pct": round(net_retention_pct, 4) if net_retention_pct is not None else None,
        "yoy_growth_pct": round(yoy_growth_pct, 4) if yoy_growth_pct is not None else None,
        "lost_only_retention_pct": round(lost_only_retention_pct, 4) if lost_only_retention_pct is not None else None,
        "punitive_retention_pct": round(punitive_retention_pct, 4) if punitive_retention_pct is not None else None,
    }

    # Top N customers
    latest_period = periods[-1]
    latest_arr = pivot[latest_period].copy()
    top_custs = latest_arr.nlargest(top_n)
    total_arr_raw = latest_arr.sum()

    top_customers = []
    for rank, (cust_id, arr_val) in enumerate(top_custs.items(), 1):
        trend = [float(pivot.loc[cust_id, p] / sf) for p in periods]
        if len(periods) >= 2:
            prev_arr = float(pivot.loc[cust_id, periods[-2]] / sf)
            curr_arr = float(arr_val / sf)
            change_pct = (curr_arr / prev_arr - 1) if prev_arr > 0 else None
        else:
            change_pct = None

        if change_pct is not None:
            if change_pct > 0.05:
                status = "Growth"
            elif change_pct < -0.05:
                status = "Declining"
            else:
                status = "Stable"
        else:
            status = "New"

        cust_str = str(cust_id)
        top_customers.append({
            "rank": rank,
            "name": cust_str,
            "arr": round(float(arr_val / sf), 2),
            "change_pct": round(change_pct, 4) if change_pct is not None else None,
            "pct_of_total": round(float(arr_val / total_arr_raw), 4) if total_arr_raw > 0 else 0,
            "trend": [round(v, 2) for v in trend],
            "status": status,
            "attributes": {k: str(v) for k, v in attr_lookup.get(cust_str, {}).items()},
            "cohort": cohort_map.get(cust_str, ""),
        })

    return {
        "periods": periods,
        "arr_over_time": [round(v, 2) for v in arr_over_time],
        "arr_growth_pcts": arr_growth_pcts,
        "waterfall": waterfall,
        "stats": stats,
        "top_customers": top_customers,
        "latest_period_label": latest_period_label,
        "latest_period_date": latest_period_date,
    }


def _compute_cohort(pivot: pd.DataFrame, periods: list[str],
                    scale_factor: int, cohort_map: dict) -> dict:
    """Compute cohort analysis data."""
    sf = scale_factor

    if not cohort_map:
        return {"periods": periods, "cohorts": []}

    cohort_series = pd.Series(cohort_map)
    unique_cohorts = [p for p in periods if p in cohort_series.values]

    cohorts = []
    for cohort_label in unique_cohorts:
        cust_in_cohort = cohort_series[cohort_series == cohort_label].index
        cohort_pivot = pivot.loc[cust_in_cohort]

        cohort_idx = periods.index(cohort_label)

        # Pre-cohort periods return None; post-cohort 0s stay as 0 (all churned)
        arr_values = [
            None if i < cohort_idx else round(float(cohort_pivot[p].sum() / sf), 2)
            for i, p in enumerate(periods)
        ]
        cust_counts = [
            None if i < cohort_idx else int((cohort_pivot[p] > 0).sum())
            for i, p in enumerate(periods)
        ]
        starting_arr = cohort_pivot[cohort_label].sum()
        starting_count = (cohort_pivot[cohort_label] > 0).sum()

        ndr = []
        for i, p in enumerate(periods):
            if i < cohort_idx or starting_arr == 0:
                ndr.append(None)
            else:
                ndr.append(round(float(cohort_pivot[p].sum() / starting_arr), 4))

        logo_ret = []
        for i, p in enumerate(periods):
            if i < cohort_idx or starting_count == 0:
                logo_ret.append(None)
            else:
                logo_ret.append(round(float((cohort_pivot[p] > 0).sum() / starting_count), 4))

        cohorts.append({
            "label": cohort_label,
            "count": int(starting_count),
            "starting_arr": round(float(starting_arr / sf), 2),
            "arr": arr_values,
            "customers": cust_counts,
            "ndr": ndr,
            "logo_retention": logo_ret,
        })

    return {
        "periods": periods,
        "cohorts": cohorts,
    }


def compute_dashboard(session_dir: str, filepath: str,
                      granularity: str | None = None,
                      filters: Optional[Dict[str, str]] = None,
                      top_n: int = 10) -> dict:
    """
    Main entry point: compute all dashboard data for a session.

    filters: dict of {attribute_display_name: value} to filter raw data.
    top_n: number of top customers to return in overview.
    """
    config = load_config(session_dir)
    raw_df = _read_raw_data(filepath, config)
    fy_month = config["fiscal_year_end_month"]
    scale_factor = config["scale_factor"]
    output_grans = config["output_granularities"]
    attr_names = list(config.get("attributes", {}).keys())

    # Compute attribute options from unfiltered data so dropdowns always show all values
    attribute_options = _compute_attribute_options(raw_df, attr_names)

    # Apply attribute filters
    if filters:
        for attr_name, attr_value in filters.items():
            if attr_name in raw_df.columns and attr_value:
                raw_df = raw_df[raw_df[attr_name] == attr_value]

    # Determine which granularity to use
    if granularity and granularity in output_grans:
        target_gran = granularity
    else:
        pref_order = ["annual", "quarterly", "monthly"]
        target_gran = next((g for g in pref_order if g in output_grans), output_grans[0])

    available = [g for g in ["annual", "quarterly", "monthly"] if g in output_grans]

    # Aggregate raw data to target granularity
    agg_df, period_date_map = _aggregate_to_granularity(raw_df, target_gran, fy_month)
    pivot, periods = _build_arr_matrix(agg_df)
    yoy_offset = _get_yoy_offset(target_gran)
    derived = _compute_derived(pivot, periods, yoy_offset)

    # Build shared lookups used by both overview and cohort
    cohort_map = _build_cohort_map(pivot, periods)
    attr_lookup = _build_attr_lookup(raw_df, attr_names)

    overview = _compute_overview(
        pivot, periods, derived, yoy_offset, scale_factor,
        period_date_map, cohort_map, attr_lookup, attr_names, top_n
    )
    cohort = _compute_cohort(pivot, periods, scale_factor, cohort_map)

    return {
        "overview": overview,
        "cohort": cohort,
        "granularity": target_gran,
        "available_granularities": available,
        "scale_factor": scale_factor,
        "attribute_options": attribute_options,
    }
