"""
Wizard API routes.

POST /api/upload           — upload Excel file, return session + sheet names
POST /api/detect-columns   — auto-detect column roles
POST /api/generate         — build config, run engine, return download token
GET  /api/download/{id}    — stream the generated Excel for download
"""

import asyncio
import json
import os
import sys

from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook

# Ensure the engine package is importable
ENGINE_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..', '..', 'data-pack-app'))
if ENGINE_PATH not in sys.path:
    sys.path.insert(0, ENGINE_PATH)

from engine.generator import generate_data_pack  # noqa: E402

from schemas import (
    UploadResponse, DetectColumnsRequest, DetectColumnsResponse,
    GenerateRequest, GenerateResponse,
    ColumnInfo, AttributeCol, DetectedMapping,
)
from temp_store import (
    create_session, get_upload_path, get_output_path, get_session_dir,
    session_exists,
)
from services.detect import detect_columns
from services.config_builder import build_engine_config

router = APIRouter(prefix="/api")


# ---------- Upload ----------

@router.post("/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...)):
    """Accept an .xlsx upload, store it, return sheet names."""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Please upload an .xlsx or .xlsm file.")

    session_id = create_session()
    upload_path = get_upload_path(session_id)

    # Stream-write to disk
    contents = await file.read()
    with open(upload_path, 'wb') as f:
        f.write(contents)

    # Read sheet names
    try:
        wb = load_workbook(upload_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        raise HTTPException(400, f"Could not read workbook: {e}")

    return UploadResponse(
        session_id=session_id,
        filename=file.filename,
        sheet_names=sheet_names,
    )


# ---------- Detect Columns ----------

@router.post("/detect-columns", response_model=DetectColumnsResponse)
async def detect(req: DetectColumnsRequest):
    """Auto-detect column roles in the selected sheet."""
    if not session_exists(req.session_id):
        raise HTTPException(404, "Session not found.")

    filepath = get_upload_path(req.session_id)
    if not os.path.isfile(filepath):
        raise HTTPException(404, "Uploaded file not found.")

    try:
        result = detect_columns(filepath, req.sheet_name)
    except Exception as e:
        raise HTTPException(400, f"Detection failed: {e}")

    mapping = result["detected_mapping"]
    return DetectColumnsResponse(
        columns=[ColumnInfo(**c) for c in result["columns"]],
        detected_mapping=DetectedMapping(
            date_col=mapping["date_col"],
            customer_id_col=mapping["customer_id_col"],
            arr_col=mapping["arr_col"],
            attribute_cols=[AttributeCol(**a) for a in mapping["attribute_cols"]],
        ),
        row_count=result["row_count"],
        auto_scale_factor=result["auto_scale_factor"],
    )


# ---------- Generate ----------

@router.post("/generate", response_model=GenerateResponse)
async def generate(req: GenerateRequest):
    """Build engine config from wizard state and generate the Excel output."""
    if not session_exists(req.session_id):
        raise HTTPException(404, "Session not found.")

    filepath = get_upload_path(req.session_id)
    output_path = get_output_path(req.session_id)

    # Build the engine config dict
    # We need row_count and scale_factor from the detect step;
    # re-detect quickly since they're lightweight reads.
    det = detect_columns(filepath, req.sheet_name)

    config = build_engine_config(
        filepath=filepath,
        session_id=req.session_id,
        sheet_name=req.sheet_name,
        data_type=req.data_type,
        date_col=req.column_mapping.date_col,
        customer_id_col=req.column_mapping.customer_id_col,
        arr_col=req.column_mapping.arr_col,
        attributes=[a.model_dump() for a in req.attributes],
        output_granularities=req.output_granularities,
        fiscal_year_end_month=req.fiscal_year_end_month,
        row_count=det["row_count"],
        scale_factor=det["auto_scale_factor"],
    )

    # Persist config for the dashboard endpoint
    config_path = os.path.join(get_session_dir(req.session_id), "config.json")
    # Convert OrderedDict to regular dict for JSON serialization
    config_serializable = dict(config)
    config_serializable["attributes"] = dict(config_serializable["attributes"])
    with open(config_path, "w") as f:
        json.dump(config_serializable, f)

    # Run the engine in a thread pool (CPU-bound openpyxl work)
    loop = asyncio.get_event_loop()
    try:
        await loop.run_in_executor(
            None, generate_data_pack, config, filepath, output_path)
    except Exception as e:
        raise HTTPException(500, f"Generation failed: {e}")

    return GenerateResponse(status="complete", download_id=req.session_id)


# ---------- Download ----------

@router.get("/download/{download_id}")
async def download(download_id: str):
    """Stream the generated Excel file for download."""
    if not session_exists(download_id):
        raise HTTPException(404, "Session not found.")

    output_path = get_output_path(download_id)
    if not os.path.isfile(output_path):
        raise HTTPException(404, "Output file not ready.")

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Data Pack Output.xlsx",
    )
